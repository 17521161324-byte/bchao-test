"""ASR 优化评估配置方案路由"""
import hashlib
import io
import json
import math
import re
from datetime import datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel, Field
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models import (
    AsrOptimizationPlan,
    AudioSeg,
    BUltraResult,
    DateFolder,
    ModelConfig,
    OptimizationFieldReviewMark,
    PatientAsrResult,
    PatientLlmResult,
    PatientRecord,
    PromptTemplate,
)


router = APIRouter()


class AsrOptimizationPlanCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=120)
    asr_model_id: int
    params: dict[str, Any] = Field(default_factory=dict)
    config_hash: str = Field(..., min_length=1, max_length=64)
    source: str = "custom"


class AsrOptimizationPlanUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=1, max_length=120)
    asr_model_id: Optional[int] = None
    params: Optional[dict[str, Any]] = None
    config_hash: Optional[str] = Field(None, min_length=1, max_length=64)


class AsrOptimizationPlanOut(BaseModel):
    id: int
    name: str
    asr_model_id: int
    asr_model_name: Optional[str] = None
    params: dict[str, Any]
    config_hash: str
    source: str
    created_at: datetime
    updated_at: datetime

    model_config = {"from_attributes": True}


class OptimizationFieldReviewMarkCreate(BaseModel):
    patient_id: int
    field_group: str = Field(..., min_length=1, max_length=50)
    field_key: Optional[str] = Field(None, max_length=50)
    asr_config_hash: Optional[str] = Field(None, max_length=64)
    asr_result_id: Optional[int] = None
    llm_result_id: int
    mark_type: str
    reason: Optional[str] = Field(None, max_length=100)
    note: Optional[str] = None


class OptimizationFieldReviewMarkOut(BaseModel):
    id: int
    patient_id: int
    field_group: str
    field_key: Optional[str] = None
    asr_config_hash: Optional[str] = None
    asr_result_id: Optional[int] = None
    llm_result_id: int
    llm_model_id: Optional[int] = None
    prompt_template_id: Optional[int] = None
    prompt_template_name: Optional[str] = None
    prompt_content_hash: Optional[str] = None
    mark_type: str
    reason: Optional[str] = None
    note: Optional[str] = None
    created_at: datetime
    updated_at: datetime

    model_config = {"from_attributes": True}


def _to_out(plan: AsrOptimizationPlan) -> AsrOptimizationPlanOut:
    return AsrOptimizationPlanOut(
        id=plan.id,
        name=plan.name,
        asr_model_id=plan.asr_model_id,
        asr_model_name=plan.asr_model.name if plan.asr_model else None,
        params=plan.params or {},
        config_hash=plan.config_hash,
        source=plan.source or "custom",
        created_at=plan.created_at,
        updated_at=plan.updated_at,
    )


def _mark_to_out(mark: OptimizationFieldReviewMark) -> OptimizationFieldReviewMarkOut:
    return OptimizationFieldReviewMarkOut(
        id=mark.id,
        patient_id=mark.patient_id,
        field_group=mark.field_group,
        field_key=mark.field_key,
        asr_config_hash=mark.asr_config_hash,
        asr_result_id=mark.asr_result_id,
        llm_result_id=mark.llm_result_id,
        llm_model_id=mark.llm_model_id,
        prompt_template_id=mark.prompt_template_id,
        prompt_template_name=mark.prompt_template_name,
        prompt_content_hash=mark.prompt_content_hash,
        mark_type=mark.mark_type,
        reason=mark.reason,
        note=mark.note,
        created_at=mark.created_at,
        updated_at=mark.updated_at,
    )


FIELD_GROUPS = [
    ("right_follicle", "右卵泡", ["right_follicle_total", "right_follicles"]),
    ("left_follicle", "左卵泡", ["left_follicle_total", "left_follicles"]),
    ("endometrium_thickness", "内膜厚度", ["endometrium_thickness"]),
    ("endometrium_type", "内膜类型", ["endometrium_type"]),
    ("right_ovary", "右卵巢", ["right_ovary_length", "right_ovary_width"]),
    ("left_ovary", "左卵巢", ["left_ovary_length", "left_ovary_width"]),
]


def _json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def _format_datetime(value: Any) -> str:
    return value.isoformat(sep=" ") if value else ""


def _norm_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if math.isnan(float(value)):
            return None
        return round(float(value), 1)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    if not match:
        return None
    return round(float(match.group()), 1)


def _normalize_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        text = value.strip()
        if text.upper() in {"N/A", "NA", "NULL", "NONE", "-"}:
            return None
        return text.upper()
    if isinstance(value, list):
        follicles = _normalize_follicles(value)
        return follicles if follicles else sorted([_normalize_value(item) for item in value], key=lambda item: _json_text(item))
    if isinstance(value, dict):
        return {k: _normalize_value(v) for k, v in value.items()}
    return value


def _normalize_follicles(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except Exception:
            return []
    if not value:
        return []
    if isinstance(value, dict):
        value = [value]
    if not isinstance(value, list):
        return []
    counts: dict[str, int] = {}
    for item in value:
        if isinstance(item, dict):
            size = _norm_number(item.get("size") or item.get("diameter") or item.get("value"))
            count = item.get("count", 1)
        else:
            size = _norm_number(item)
            count = 1
        try:
            count = int(float(count or 1))
        except Exception:
            count = 1
        if size is None or count <= 0:
            continue
        key = f"{size:.1f}"
        counts[key] = counts.get(key, 0) + count
    return [{"size": key, "count": counts[key]} for key in sorted(counts, key=lambda item: float(item))]


def _follicles_to_str(value: Any) -> str:
    follicles = _normalize_follicles(value)
    return "、".join(f"{item['size']}×{item['count']}" for item in follicles) if follicles else ""


def _value_to_str(value: Any) -> str:
    if value is None:
        return ""
    follicles = _normalize_follicles(value)
    if follicles:
        return _follicles_to_str(follicles)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _get_gt_value(gt: BUltraResult | None, field_key: str) -> Any:
    return getattr(gt, field_key, None) if gt else None


def _compare_field(structured: dict[str, Any], gt: BUltraResult | None, field_key: str) -> tuple[str, str, str]:
    llm_value = structured.get(field_key) if structured else None
    gt_value = _get_gt_value(gt, field_key)
    llm_norm = _normalize_value(llm_value)
    gt_norm = _normalize_value(gt_value)
    if llm_norm is None and gt_norm is None:
        status = "匹配"
    elif llm_norm == gt_norm:
        status = "匹配"
    else:
        status = "不匹配"
    return status, _value_to_str(llm_value), _value_to_str(gt_value)


def _compare_group(structured: dict[str, Any], gt: BUltraResult | None, field_keys: list[str]) -> tuple[str, str, str]:
    statuses: list[str] = []
    llm_parts: list[str] = []
    gt_parts: list[str] = []
    for key in field_keys:
        status, llm_text, gt_text = _compare_field(structured, gt, key)
        statuses.append(status)
        llm_parts.append(f"{key}={llm_text}")
        gt_parts.append(f"{key}={gt_text}")
    return ("匹配" if all(item == "匹配" for item in statuses) else "不匹配", "；".join(llm_parts), "；".join(gt_parts))


def _apply_sheet_style(ws):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_wrap = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = header_wrap if row_idx == 1 else body_wrap
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 48)


def _append_rows(ws, headers: list[str], rows: list[list[Any]]):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _apply_sheet_style(ws)


async def _ensure_asr_model(db: AsyncSession, model_id: int) -> None:
    result = await db.execute(
        select(ModelConfig).where(ModelConfig.id == model_id, ModelConfig.model_type == "asr")
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="ASR 模型不存在")


@router.get("/plans", response_model=list[AsrOptimizationPlanOut])
async def list_plans(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(AsrOptimizationPlan)
        .options(selectinload(AsrOptimizationPlan.asr_model))
        .order_by(AsrOptimizationPlan.updated_at.desc(), AsrOptimizationPlan.id.desc())
    )
    return [_to_out(plan) for plan in result.scalars().all()]


@router.post("/plans", response_model=AsrOptimizationPlanOut)
async def save_plan(data: AsrOptimizationPlanCreate, db: AsyncSession = Depends(get_db)):
    await _ensure_asr_model(db, data.asr_model_id)
    config_hash = data.config_hash.strip()[:64]
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="方案名称不能为空")

    result = await db.execute(
        select(AsrOptimizationPlan).where(AsrOptimizationPlan.config_hash == config_hash)
    )
    plan = result.scalar_one_or_none()
    if plan:
        plan.name = name
        plan.asr_model_id = data.asr_model_id
        plan.params = data.params or {}
        plan.source = "custom"
    else:
        plan = AsrOptimizationPlan(
            name=name,
            asr_model_id=data.asr_model_id,
            params=data.params or {},
            config_hash=config_hash,
            source="custom",
        )
        db.add(plan)

    await db.commit()
    await db.refresh(plan)
    await db.refresh(plan, attribute_names=["asr_model"])
    return _to_out(plan)


@router.put("/plans/{plan_id}", response_model=AsrOptimizationPlanOut)
async def update_plan(
    plan_id: int,
    data: AsrOptimizationPlanUpdate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(AsrOptimizationPlan).where(AsrOptimizationPlan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="方案不存在")
    if (plan.source or "custom") != "custom":
        raise HTTPException(status_code=400, detail="仅支持修改自定义方案")

    update_data = data.model_dump(exclude_unset=True)
    if "name" in update_data:
        name = (update_data["name"] or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="方案名称不能为空")
        plan.name = name
    if "asr_model_id" in update_data and update_data["asr_model_id"] is not None:
        await _ensure_asr_model(db, update_data["asr_model_id"])
        plan.asr_model_id = update_data["asr_model_id"]
    if "params" in update_data and update_data["params"] is not None:
        plan.params = update_data["params"]
    if "config_hash" in update_data and update_data["config_hash"]:
        plan.config_hash = update_data["config_hash"].strip()[:64]

    await db.commit()
    await db.refresh(plan)
    await db.refresh(plan, attribute_names=["asr_model"])
    return _to_out(plan)


@router.delete("/plans/{plan_id}")
async def delete_plan(plan_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(AsrOptimizationPlan).where(AsrOptimizationPlan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="方案不存在")
    if (plan.source or "custom") != "custom":
        raise HTTPException(status_code=400, detail="仅支持删除自定义方案")

    await db.delete(plan)
    await db.commit()
    return {"message": "删除成功"}


@router.get("/field-review-marks", response_model=list[OptimizationFieldReviewMarkOut])
async def list_field_review_marks(
    llm_result_ids: str = Query("", description="逗号分隔的 LLM 结果 ID"),
    db: AsyncSession = Depends(get_db),
):
    ids = [int(item) for item in llm_result_ids.split(",") if item.strip().isdigit()]
    if not ids:
        return []
    result = await db.execute(
        select(OptimizationFieldReviewMark)
        .where(OptimizationFieldReviewMark.llm_result_id.in_(ids))
        .order_by(OptimizationFieldReviewMark.updated_at.desc(), OptimizationFieldReviewMark.id.desc())
    )
    return [_mark_to_out(mark) for mark in result.scalars().all()]


@router.post("/field-review-marks", response_model=OptimizationFieldReviewMarkOut)
async def upsert_field_review_mark(
    data: OptimizationFieldReviewMarkCreate,
    db: AsyncSession = Depends(get_db),
):
    if data.mark_type not in ("exclude", "mismatch_note"):
        raise HTTPException(status_code=400, detail="mark_type 必须为 exclude 或 mismatch_note")

    llm_result = await db.get(PatientLlmResult, data.llm_result_id)
    if not llm_result:
        raise HTTPException(status_code=404, detail="LLM 结果不存在")
    if llm_result.patient_id != data.patient_id:
        raise HTTPException(status_code=400, detail="LLM 结果与检查记录不匹配")

    asr_result = None
    if llm_result.asr_result_id:
        asr_result = await db.get(PatientAsrResult, llm_result.asr_result_id)

    prompt_content = llm_result.prompt_content or ""
    prompt_hash = hashlib.sha256(prompt_content.encode("utf-8")).hexdigest()[:64] if prompt_content else None
    field_group = data.field_group.strip()
    field_key = (data.field_key or "").strip() or None

    query = select(OptimizationFieldReviewMark).where(
        OptimizationFieldReviewMark.llm_result_id == data.llm_result_id,
        OptimizationFieldReviewMark.field_group == field_group,
    )
    if field_key:
        query = query.where(OptimizationFieldReviewMark.field_key == field_key)
    else:
        query = query.where(OptimizationFieldReviewMark.field_key.is_(None))

    result = await db.execute(query)
    mark = result.scalar_one_or_none()
    values = {
        "patient_id": data.patient_id,
        "field_group": field_group,
        "field_key": field_key,
        "asr_config_hash": data.asr_config_hash or (asr_result.config_hash if asr_result else None),
        "asr_result_id": data.asr_result_id or llm_result.asr_result_id,
        "llm_result_id": data.llm_result_id,
        "llm_model_id": llm_result.llm_model_id,
        "prompt_template_id": llm_result.prompt_template_id,
        "prompt_template_name": llm_result.prompt_template_name,
        "prompt_content_hash": prompt_hash,
        "mark_type": data.mark_type,
        "reason": data.reason or None,
        "note": data.note or None,
    }
    if mark:
        for key, value in values.items():
            setattr(mark, key, value)
    else:
        mark = OptimizationFieldReviewMark(**values)
        db.add(mark)

    await db.commit()
    await db.refresh(mark)
    return _mark_to_out(mark)


@router.delete("/field-review-marks")
async def delete_field_review_mark(
    llm_result_id: int = Query(...),
    field_group: str = Query(...),
    field_key: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = delete(OptimizationFieldReviewMark).where(
        OptimizationFieldReviewMark.llm_result_id == llm_result_id,
        OptimizationFieldReviewMark.field_group == field_group,
    )
    if field_key:
        stmt = stmt.where(OptimizationFieldReviewMark.field_key == field_key)
    else:
        stmt = stmt.where(OptimizationFieldReviewMark.field_key.is_(None))

    result = await db.execute(stmt.execution_options(synchronize_session=False))
    await db.commit()
    return {"message": "已清除标记", "deleted": result.rowcount or 0}


@router.delete("/plans/by-hash/{config_hash}")
async def delete_plan_by_hash(config_hash: str, db: AsyncSession = Depends(get_db)):
    """按配置指纹删除优化评估方案和关联历史结果。

    业务含义:
    - 删除 asr_optimization_plans 中同 config_hash 的保存配置；
    - 删除 patient_asr_results 中 source=asr_optimization 且同 config_hash 的历史 ASR；
    - 删除这些 ASR 结果关联的 source=asr_optimization LLM 结果；
    - 不删除普通 ASR/LLM 结果、不删除模型配置、不删除患者数据。
    """
    target_hash = (config_hash or "").strip()[:64]
    if not target_hash:
        raise HTTPException(status_code=400, detail="配置指纹不能为空")

    asr_result = await db.execute(
        select(PatientAsrResult.id).where(
            PatientAsrResult.source == "asr_optimization",
            PatientAsrResult.config_hash == target_hash,
        )
    )
    asr_ids = [row[0] for row in asr_result.fetchall()]

    deleted_llm = 0
    deleted_marks = 0
    if asr_ids:
        llm_ids_result = await db.execute(
            select(PatientLlmResult.id).where(
                PatientLlmResult.source == "asr_optimization",
                PatientLlmResult.asr_result_id.in_(asr_ids),
            )
        )
        llm_ids = [row[0] for row in llm_ids_result.fetchall()]
        if llm_ids:
            mark_result = await db.execute(
                delete(OptimizationFieldReviewMark)
                .where(OptimizationFieldReviewMark.llm_result_id.in_(llm_ids))
                .execution_options(synchronize_session=False)
            )
            deleted_marks = mark_result.rowcount or 0
        llm_result = await db.execute(
            delete(PatientLlmResult)
            .where(
                PatientLlmResult.source == "asr_optimization",
                PatientLlmResult.asr_result_id.in_(asr_ids),
            )
            .execution_options(synchronize_session=False)
        )
        deleted_llm = llm_result.rowcount or 0

    asr_delete_result = await db.execute(
        delete(PatientAsrResult)
        .where(
            PatientAsrResult.source == "asr_optimization",
            PatientAsrResult.config_hash == target_hash,
        )
        .execution_options(synchronize_session=False)
    )
    plan_delete_result = await db.execute(
        delete(AsrOptimizationPlan)
        .where(AsrOptimizationPlan.config_hash == target_hash)
        .execution_options(synchronize_session=False)
    )

    await db.commit()
    return {
        "message": "删除成功",
        "config_hash": target_hash,
        "deleted_plans": plan_delete_result.rowcount or 0,
        "deleted_asr_results": asr_delete_result.rowcount or 0,
        "deleted_llm_results": deleted_llm,
        "deleted_field_review_marks": deleted_marks,
    }


@router.post("/export-full")
async def export_full_optimization_data(
    body: dict[str, Any],
    db: AsyncSession = Depends(get_db),
):
    """导出当前 ASR 配置指纹下的完整优化评估数据。"""
    from openpyxl import Workbook

    config_hash = str(body.get("config_hash") or "").strip()
    if not config_hash:
        raise HTTPException(status_code=400, detail="缺少 config_hash")
    dates = [str(item).strip() for item in (body.get("dates") or []) if str(item).strip()]
    hash_prefix = config_hash[:16]

    asr_stmt = (
        select(PatientAsrResult)
        .where(PatientAsrResult.config_hash == config_hash)
        .order_by(PatientAsrResult.date.asc(), PatientAsrResult.record_id.asc(), PatientAsrResult.created_at.desc())
    )
    if dates:
        asr_stmt = asr_stmt.where(PatientAsrResult.date.in_(dates))
    asr_rows = (await db.execute(asr_stmt)).scalars().all()
    if not asr_rows:
        raise HTTPException(status_code=404, detail="当前指纹/筛选日期下暂无 ASR 结果")

    patient_ids = sorted({row.patient_id for row in asr_rows})
    asr_ids = [row.id for row in asr_rows]

    patients = (await db.execute(
        select(PatientRecord)
        .options(selectinload(PatientRecord.date_folder), selectinload(PatientRecord.segs))
        .where(PatientRecord.id.in_(patient_ids))
    )).scalars().all()
    patient_map = {row.id: row for row in patients}

    gt_rows = (await db.execute(select(BUltraResult).where(BUltraResult.patient_id.in_(patient_ids)))).scalars().all()
    gt_map = {row.patient_id: row for row in gt_rows}

    llm_rows = (await db.execute(
        select(PatientLlmResult)
        .where(PatientLlmResult.asr_result_id.in_(asr_ids))
        .order_by(PatientLlmResult.patient_id.asc(), PatientLlmResult.created_at.asc(), PatientLlmResult.id.asc())
    )).scalars().all()
    llm_ids = [row.id for row in llm_rows]

    mark_rows: list[OptimizationFieldReviewMark] = []
    if llm_ids:
        mark_rows = (await db.execute(
            select(OptimizationFieldReviewMark)
            .where(OptimizationFieldReviewMark.llm_result_id.in_(llm_ids))
            .order_by(OptimizationFieldReviewMark.patient_id.asc(), OptimizationFieldReviewMark.field_group.asc())
        )).scalars().all()
    marks_by_llm_group = {(row.llm_result_id, row.field_group): row for row in mark_rows}

    prompt_ids = sorted({row.prompt_template_id for row in llm_rows if row.prompt_template_id})
    prompt_rows = []
    if prompt_ids:
        prompt_rows = (await db.execute(select(PromptTemplate).where(PromptTemplate.id.in_(prompt_ids)))).scalars().all()
    prompt_map = {row.id: row for row in prompt_rows}

    seg_rows = (await db.execute(
        select(AudioSeg)
        .where(AudioSeg.patient_id.in_(patient_ids))
        .order_by(AudioSeg.patient_id.asc(), AudioSeg.seg_index.asc())
    )).scalars().all()

    plan = (await db.execute(select(AsrOptimizationPlan).where(AsrOptimizationPlan.config_hash == config_hash))).scalar_one_or_none()

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 汇总
    ws = wb.create_sheet("汇总统计")
    combo_counter: dict[tuple[str, str], dict[str, Any]] = {}
    field_stats: dict[str, dict[str, int]] = {group: {"total": 0, "matched": 0, "excluded": 0, "abnormal": 0} for group, _, _ in FIELD_GROUPS}
    for llm in llm_rows:
        key = (llm.prompt_template_name or f"提示词{llm.prompt_template_id or ''}", llm.llm_model_name or f"LLM{llm.llm_model_id or ''}")
        combo_counter.setdefault(key, {"count": 0, "matched": 0, "total": 0})
        combo_counter[key]["count"] += 1
        gt = gt_map.get(llm.patient_id)
        structured = llm.structured_result or {}
        for group_key, _, field_keys in FIELD_GROUPS:
            mark = marks_by_llm_group.get((llm.id, group_key))
            if mark and mark.mark_type == "exclude":
                field_stats[group_key]["excluded"] += 1
                continue
            if mark and mark.mark_type == "mismatch_note":
                field_stats[group_key]["abnormal"] += 1
            status, _, _ = _compare_group(structured, gt, field_keys)
            field_stats[group_key]["total"] += 1
            combo_counter[key]["total"] += 1
            if status == "匹配":
                field_stats[group_key]["matched"] += 1
                combo_counter[key]["matched"] += 1

    summary_rows = [
        ["配置指纹", config_hash],
        ["方案名称", plan.name if plan else ""],
        ["日期范围", "、".join(dates) if dates else "全部"],
        ["检查记录数", len(patient_ids)],
        ["ASR结果数", len(asr_rows)],
        ["LLM结果数", len(llm_rows)],
        ["人工标记数", len(mark_rows)],
        ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    _append_rows(ws, ["项目", "值"], summary_rows)

    ws = wb.create_sheet("组合成功率")
    combo_rows = []
    for (prompt_name, llm_name), stat in sorted(combo_counter.items(), key=lambda item: (item[0][0], item[0][1])):
        total = stat["total"]
        combo_rows.append([prompt_name, llm_name, stat["count"], stat["matched"], total, f"{(stat['matched'] / total * 100):.1f}%" if total else ""])
    _append_rows(ws, ["提示词模板", "LLM模型", "结果条数", "匹配字段数", "参与字段数", "字段成功率"], combo_rows)

    ws = wb.create_sheet("字段成功率")
    field_rows = []
    for group_key, label, _ in FIELD_GROUPS:
        stat = field_stats[group_key]
        total = stat["total"]
        field_rows.append([label, stat["matched"], total, f"{(stat['matched'] / total * 100):.1f}%" if total else "", stat["excluded"], stat["abnormal"]])
    _append_rows(ws, ["字段", "匹配数", "参与数", "成功率", "人工排除", "人工异常"], field_rows)

    # Sheet 检查 + ASR
    ws = wb.create_sheet("检查_ASR明细")
    asr_detail_rows = []
    for asr in asr_rows:
        patient = patient_map.get(asr.patient_id)
        gt = gt_map.get(asr.patient_id)
        asr_detail_rows.append([
            asr.patient_id,
            asr.record_id,
            asr.date,
            patient.note if patient else "",
            len(patient.segs) if patient else 0,
            "有" if gt else "无",
            asr.id,
            asr.asr_model_id,
            asr.asr_model_name,
            asr.provider,
            asr.source,
            asr.experiment_key,
            asr.config_hash,
            _json_text(asr.config_snapshot),
            asr.status,
            asr.error_message or "",
            len(asr.full_transcript or ""),
            asr.full_transcript or "",
            _format_datetime(asr.created_at),
        ])
    _append_rows(ws, [
        "检查记录ID", "病历号", "日期", "检查备注", "录音段数", "真实B超", "ASR结果ID", "ASR模型ID", "ASR模型名称",
        "Provider", "来源", "实验Key", "配置指纹", "配置快照", "状态", "错误", "转写字数", "完整转写", "创建时间",
    ], asr_detail_rows)

    # Sheet LLM
    ws = wb.create_sheet("LLM结果明细")
    llm_detail_rows = []
    asr_map = {row.id: row for row in asr_rows}
    for llm in llm_rows:
        asr = asr_map.get(llm.asr_result_id)
        gt = gt_map.get(llm.patient_id)
        structured = llm.structured_result or {}
        llm_detail_rows.append([
            llm.id,
            llm.patient_id,
            asr.record_id if asr else "",
            asr.date if asr else "",
            llm.asr_result_id,
            asr.asr_model_name if asr else "",
            llm.llm_model_id,
            llm.llm_model_name,
            llm.prompt_template_id,
            llm.prompt_template_name,
            len(llm.prompt_content or ""),
            llm.status,
            llm.accuracy,
            llm.error_message or "",
            structured.get("right_follicle_total"),
            _follicles_to_str(structured.get("right_follicles")),
            structured.get("left_follicle_total"),
            _follicles_to_str(structured.get("left_follicles")),
            structured.get("endometrium_thickness"),
            structured.get("endometrium_type"),
            structured.get("right_ovary_length"),
            structured.get("right_ovary_width"),
            structured.get("left_ovary_length"),
            structured.get("left_ovary_width"),
            structured.get("remark"),
            llm.summary_text or "",
            llm.raw_output or "",
            _json_text(llm.structured_result),
            _format_datetime(llm.created_at),
            "有" if gt else "无",
        ])
    _append_rows(ws, [
        "LLM结果ID", "检查记录ID", "病历号", "日期", "ASR结果ID", "ASR模型", "LLM模型ID", "LLM模型",
        "提示词模板ID", "提示词模板", "提示词长度", "状态", "原准确率", "错误",
        "右卵泡总数", "右卵泡明细", "左卵泡总数", "左卵泡明细", "内膜厚度", "内膜类型",
        "右卵巢长", "右卵巢宽", "左卵巢长", "左卵巢宽", "备注", "LLM总结", "LLM原始输出",
        "结构化JSON", "创建时间", "真实B超",
    ], llm_detail_rows)

    # Sheet 字段对比
    ws = wb.create_sheet("字段对比明细")
    compare_rows = []
    for llm in llm_rows:
        asr = asr_map.get(llm.asr_result_id)
        gt = gt_map.get(llm.patient_id)
        structured = llm.structured_result or {}
        for group_key, label, field_keys in FIELD_GROUPS:
            status, llm_text, gt_text = _compare_group(structured, gt, field_keys)
            mark = marks_by_llm_group.get((llm.id, group_key))
            display_status = "排除" if mark and mark.mark_type == "exclude" else ("异常" if mark and mark.mark_type == "mismatch_note" else status)
            compare_rows.append([
                llm.patient_id,
                asr.record_id if asr else "",
                asr.date if asr else "",
                asr.config_hash if asr else "",
                llm.id,
                llm.llm_model_name,
                llm.prompt_template_name,
                label,
                display_status,
                status,
                llm_text,
                gt_text,
                mark.mark_type if mark else "",
                mark.reason if mark else "",
                mark.note if mark else "",
            ])
    _append_rows(ws, [
        "检查记录ID", "病历号", "日期", "ASR指纹", "LLM结果ID", "LLM模型", "提示词模板",
        "字段", "显示状态", "原始匹配", "LLM提取值", "真实值", "人工标记", "人工原因", "人工备注",
    ], compare_rows)

    # Sheet 人工标记
    ws = wb.create_sheet("人工标记明细")
    mark_detail_rows = []
    for mark in mark_rows:
        llm = next((row for row in llm_rows if row.id == mark.llm_result_id), None)
        asr = asr_map.get(llm.asr_result_id) if llm else None
        mark_detail_rows.append([
            mark.id,
            mark.patient_id,
            asr.record_id if asr else "",
            asr.date if asr else "",
            mark.field_group,
            mark.field_key or "",
            mark.mark_type,
            mark.reason or "",
            mark.note or "",
            mark.asr_config_hash or "",
            mark.asr_result_id or "",
            mark.llm_result_id,
            mark.llm_model_id or "",
            mark.prompt_template_id or "",
            mark.prompt_template_name or "",
            _format_datetime(mark.created_at),
            _format_datetime(mark.updated_at),
        ])
    _append_rows(ws, [
        "标记ID", "检查记录ID", "病历号", "日期", "字段组", "字段Key", "标记类型", "原因", "备注",
        "ASR指纹", "ASR结果ID", "LLM结果ID", "LLM模型ID", "提示词模板ID", "提示词模板", "创建时间", "更新时间",
    ], mark_detail_rows)

    # Sheet ASR分段
    ws = wb.create_sheet("ASR分段明细")
    seg_detail_rows = []
    for asr in asr_rows:
        segments = asr.segments or []
        if not segments:
            seg_detail_rows.append([asr.patient_id, asr.record_id, asr.date, asr.id, "", "", "", "", ""])
            continue
        for seg in segments:
            seg_detail_rows.append([
                asr.patient_id,
                asr.record_id,
                asr.date,
                asr.id,
                seg.get("seg_index") if isinstance(seg, dict) else "",
                seg.get("filename") if isinstance(seg, dict) else "",
                seg.get("duration") if isinstance(seg, dict) else "",
                seg.get("text") if isinstance(seg, dict) else _json_text(seg),
                seg.get("error") if isinstance(seg, dict) else "",
            ])
    _append_rows(ws, ["检查记录ID", "病历号", "日期", "ASR结果ID", "段号", "文件名", "时长", "分段转写", "错误"], seg_detail_rows)

    # Sheet 录音文件
    ws = wb.create_sheet("录音文件")
    audio_rows = [[seg.patient_id, patient_map.get(seg.patient_id).record_id if patient_map.get(seg.patient_id) else "", patient_map.get(seg.patient_id).date_folder.date if patient_map.get(seg.patient_id) and patient_map.get(seg.patient_id).date_folder else "", seg.seg_index, seg.filename, seg.duration, seg.file_size, seg.file_path] for seg in seg_rows]
    _append_rows(ws, ["检查记录ID", "病历号", "日期", "段号", "文件名", "时长", "大小", "路径"], audio_rows)

    # Sheet 提示词
    ws = wb.create_sheet("提示词内容")
    prompt_detail_rows = []
    seen_prompt_ids: set[int] = set()
    for llm in llm_rows:
        prompt_id = llm.prompt_template_id or 0
        if prompt_id in seen_prompt_ids:
            continue
        seen_prompt_ids.add(prompt_id)
        template = prompt_map.get(prompt_id)
        content = template.content if template else llm.prompt_content
        prompt_detail_rows.append([prompt_id or "", llm.prompt_template_name or (template.name if template else ""), len(content or ""), content or ""])
    _append_rows(ws, ["提示词模板ID", "提示词模板名称", "内容长度", "提示词内容"], prompt_detail_rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    date_part = "-".join(dates) if dates else "all"
    filename = f"ASR优化评估_{hash_prefix}_{date_part}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from urllib.parse import quote

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
