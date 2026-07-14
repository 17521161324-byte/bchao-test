"""
录音管理路由
"""
import io
import os
from fastapi import APIRouter, Depends, HTTPException, Body
from fastapi.responses import FileResponse, Response
from sqlalchemy import select, func, delete as sa_delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from loguru import logger

from app.database import get_db
from app.models import DateFolder, PatientRecord, AudioSeg, BUltraResult, PatientAsrResult, PatientLlmResult, PromptTemplate
from app.schemas import DateFolderOut, PatientRecordOut, DataStatusOut
from app.config import settings
from app.services.parser import evaluate_result, normalize_structured_result

router = APIRouter()


def _recordings_root() -> str:
    return os.path.realpath(settings.RECORDINGS_DIR)


def _path_inside_recordings(full_path: str, recordings_dir: str) -> bool:
    try:
        return os.path.commonpath([recordings_dir, full_path]) == recordings_dir
    except ValueError:
        return False


def _legacy_recordings_relative_path(raw_path: str) -> str | None:
    normalized = raw_path.replace("\\", "/")
    marker = "/recordings/"
    if marker in normalized:
        return normalized.split(marker, 1)[1]
    if normalized.startswith("recordings/"):
        return normalized.split("/", 1)[1]
    return None


def _resolve_audio_file_path(raw_path: str) -> str:
    recordings_dir = _recordings_root()
    full_path = os.path.realpath(raw_path)

    if not _path_inside_recordings(full_path, recordings_dir):
        relative_path = _legacy_recordings_relative_path(raw_path)
        if relative_path:
            full_path = os.path.realpath(os.path.join(recordings_dir, *relative_path.split("/")))

    if not _path_inside_recordings(full_path, recordings_dir):
        raise HTTPException(status_code=403, detail="禁止访问")

    if not os.path.isfile(full_path):
        raise HTTPException(status_code=404, detail="文件不存在")

    return full_path


@router.get("/tree")
async def get_audio_tree(db: AsyncSession = Depends(get_db)):
    """获取录音文件树（日期→病历号→seg）"""
    result = await db.execute(
        select(DateFolder)
        .options(
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.segs),
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.result),
        )
        .order_by(DateFolder.date.desc())
    )
    folders = result.unique().scalars().all()

    # 构建响应
    output = []
    for folder in folders:
        patients_out = []
        for p in folder.patients:
            has_result = p.result is not None
            patients_out.append({
                "id": p.id,
                "record_id": p.record_id,
                "date_folder_id": p.date_folder_id,
                "timestamp_folder": p.timestamp_folder,
                "note": p.note or "",
                "segs": [
                    {
                        "id": s.id,
                        "seg_index": s.seg_index,
                        "filename": s.filename,
                        "duration": s.duration,
                        "file_path": s.file_path,
                        "file_size": s.file_size,
                    }
                    for s in sorted(p.segs, key=lambda x: x.seg_index)
                ],
                "has_result": has_result,
            })
        output.append({
            "id": folder.id,
            "date": folder.date,
            "patient_count": len(patients_out),
            "patients": patients_out,
        })
    return output


@router.get("/file")
async def get_audio_file(path: str):
    """获取音频文件流（支持 range 请求）"""
    full_path = _resolve_audio_file_path(path)
    return FileResponse(
        full_path,
        media_type="audio/wav",
        filename=os.path.basename(full_path),
    )


@router.get("/file/{seg_id}")
async def get_audio_file_by_seg(seg_id: int, db: AsyncSession = Depends(get_db)):
    """按录音分段 ID 获取音频文件，兼容历史 Windows 绝对路径。"""
    seg = await db.get(AudioSeg, seg_id)
    if not seg:
        raise HTTPException(status_code=404, detail="录音分段不存在")

    full_path = _resolve_audio_file_path(seg.file_path or "")
    return FileResponse(
        full_path,
        media_type="audio/wav",
        filename=os.path.basename(full_path),
    )


@router.get("/status", response_model=DataStatusOut)
async def get_data_status(db: AsyncSession = Depends(get_db)):
    """获取数据匹配状态统计"""
    # 总日期数
    date_count = await db.execute(select(func.count(DateFolder.id)))
    total_dates = date_count.scalar()

    # 总病历数
    patient_count = await db.execute(select(func.count(PatientRecord.id)))
    total_patients = patient_count.scalar()

    # 已匹配数（有结果）
    matched = await db.execute(
        select(func.count(PatientRecord.id))
        .join(BUltraResult, BUltraResult.patient_id == PatientRecord.id)
    )
    matched_count = matched.scalar()

    # 仅有录音（无结果）
    audio_only = total_patients - matched_count

    # 仅有结果（无录音）- 理论上不应存在
    result_only = 0

    return {
        "total_dates": total_dates,
        "total_patients": total_patients,
        "matched_count": matched_count,
        "audio_only_count": audio_only,
        "result_only_count": result_only,
    }


@router.post("/scan")
async def scan_recordings(db: AsyncSession = Depends(get_db)):
    """
    扫描录音目录，同步文件结构到数据库
    读取 settings.RECORDINGS_DIR 下的日期→病历号→seg 层级
    """
    recordings_dir = settings.RECORDINGS_DIR
    if not os.path.isdir(recordings_dir):
        raise HTTPException(status_code=400, detail=f"录音目录不存在: {recordings_dir}")

    scanned = {"dates": 0, "patients": 0, "segs": 0}

    # 遍历日期文件夹
    for date_name in sorted(os.listdir(recordings_dir)):
        date_path = os.path.join(recordings_dir, date_name)
        if not os.path.isdir(date_path) or not date_name.isdigit():
            continue

        # 查找或创建日期记录
        result = await db.execute(select(DateFolder).where(DateFolder.date == date_name))
        date_folder = result.scalar_one_or_none()
        if not date_folder:
            date_folder = DateFolder(date=date_name, path=date_path)
            db.add(date_folder)
            await db.flush()
            scanned["dates"] += 1

        # 遍历病历号文件夹
        for record_name in sorted(os.listdir(date_path)):
            record_path = os.path.join(date_path, record_name)
            if not os.path.isdir(record_path):
                continue

            # 查找或创建病历记录
            result = await db.execute(
                select(PatientRecord).where(
                    PatientRecord.record_id == record_name,
                    PatientRecord.date_folder_id == date_folder.id,
                )
            )
            patient = result.scalar_one_or_none()
            if not patient:
                patient = PatientRecord(
                    record_id=record_name,
                    date_folder_id=date_folder.id,
                    timestamp_folder="",
                )
                db.add(patient)
                await db.flush()
                scanned["patients"] += 1

            # 查找时间戳文件夹 和 audio/ 子目录
            audio_dir = self_find_audio_dir(record_path)
            if audio_dir and patient:
                patient.timestamp_folder = os.path.basename(os.path.dirname(audio_dir))
                # 遍历 seg 文件
                existing_segs = {s.filename for s in patient.segs}
                for seg_file in sorted(os.listdir(audio_dir)):
                    if not seg_file.endswith(".wav"):
                        continue
                    if seg_file in existing_segs:
                        continue
                    seg_path = os.path.join(audio_dir, seg_file)
                    seg_index = self_parse_seg_index(seg_file)
                    seg = AudioSeg(
                        patient_id=patient.id,
                        seg_index=seg_index,
                        filename=seg_file,
                        file_path=seg_path,
                        file_size=os.path.getsize(seg_path),
                    )
                    db.add(seg)
                    scanned["segs"] += 1

    await db.commit()
    logger.info(f"扫描完成: {scanned}")
    return {"message": "扫描完成", "scanned": scanned}


@router.get("/batches")
async def get_batches(db: AsyncSession = Depends(get_db)):
    """获取批次列表（按日期）及统计"""
    result = await db.execute(
        select(DateFolder.id, DateFolder.date)
        .order_by(DateFolder.date.desc())
    )
    rows = result.all()

    batches = []
    for row in rows:
        # Count patients in this batch
        patient_count = await db.execute(
            select(func.count(PatientRecord.id))
            .where(PatientRecord.date_folder_id == row.id)
        )
        # Count matched (with results)
        matched = await db.execute(
            select(func.count(PatientRecord.id))
            .join(BUltraResult, BUltraResult.patient_id == PatientRecord.id)
            .where(PatientRecord.date_folder_id == row.id)
        )
        batches.append({
            "id": row.id,
            "date": row.date,
            "patient_count": patient_count.scalar(),
            "matched_count": matched.scalar(),
        })
    return batches


@router.get("/verify")
async def verify_data(date: str = None, db: AsyncSession = Depends(get_db)):
    """数据核对：检测异常数据"""
    import os

    result = await db.execute(
        select(DateFolder)
        .options(
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.segs),
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.result),
        )
        .order_by(DateFolder.date.desc())
    )
    folders = result.unique().scalars().all()

    if date:
        folders = [f for f in folders if f.date == date]

    issues = []
    for folder in folders:
        # Detect duplicates
        seen_ids = {}
        for p in folder.patients:
            if p.record_id in seen_ids:
                issues.append({
                    "type": "duplicate",
                    "date": folder.date,
                    "record_id": p.record_id,
                    "patient_id": p.id,
                    "detail": f"重复病历号 (同日期已有 id={seen_ids[p.record_id]})",
                    "action": "merge_or_delete",
                })
            else:
                seen_ids[p.record_id] = p.id

        # Check each patient
        for p in folder.patients:
            has_segs = len(p.segs) > 0
            has_result = p.result is not None

            # Check if files actually exist
            missing_files = []
            for s in p.segs:
                try:
                    _resolve_audio_file_path(s.file_path or "")
                except HTTPException:
                    missing_files.append(s.filename)

            if not has_segs and has_result:
                issues.append({
                    "type": "no_audio_has_result",
                    "date": folder.date,
                    "record_id": p.record_id,
                    "patient_id": p.id,
                    "detail": f"无录音但有B超结果",
                    "action": "delete_result",
                })
            elif has_segs and missing_files:
                issues.append({
                    "type": "missing_files",
                    "date": folder.date,
                    "record_id": p.record_id,
                    "patient_id": p.id,
                    "detail": f"缺少 {len(missing_files)} 个录音文件",
                    "action": "investigate",
                    "missing_files": missing_files[:5],
                })

    return {
        "total_issues": len(issues),
        "issues": issues,
    }


@router.post("/records/export-latest")
async def export_latest_llm_results(
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
):
    """导出数据管理当前记录范围内每个检查记录的最新一次 LLM 提取结果。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    raw_patient_ids = data.get("patient_ids") or []
    try:
        patient_ids = [int(pid) for pid in raw_patient_ids]
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="patient_ids 参数格式错误")

    # 去重但保留前端当前表格顺序
    patient_ids = list(dict.fromkeys(patient_ids))
    if not patient_ids:
        raise HTTPException(status_code=400, detail="请选择需要导出的检查记录")

    patient_result = await db.execute(
        select(PatientRecord)
        .options(
            selectinload(PatientRecord.date_folder),
            selectinload(PatientRecord.result),
        )
        .where(PatientRecord.id.in_(patient_ids))
    )
    patient_map = {p.id: p for p in patient_result.scalars().all()}
    patients = [patient_map[pid] for pid in patient_ids if pid in patient_map]
    if not patients:
        raise HTTPException(status_code=404, detail="未找到可导出的检查记录")

    llm_result = await db.execute(
        select(PatientLlmResult)
        .where(PatientLlmResult.patient_id.in_([p.id for p in patients]))
        .order_by(
            PatientLlmResult.patient_id.asc(),
            PatientLlmResult.created_at.desc(),
            PatientLlmResult.id.desc(),
        )
    )
    latest_llm_by_patient = {}
    llm_rows = llm_result.scalars().all()
    for llm in llm_rows:
        if llm.patient_id not in latest_llm_by_patient:
            latest_llm_by_patient[llm.patient_id] = llm

    asr_ids = list({llm.asr_result_id for llm in llm_rows if llm.asr_result_id})
    asr_map = {}
    if asr_ids:
        asr_result = await db.execute(
            select(PatientAsrResult).where(PatientAsrResult.id.in_(asr_ids))
        )
        asr_map = {a.id: a for a in asr_result.scalars().all()}

    tmpl_ids = list({llm.prompt_template_id for llm in llm_rows if llm.prompt_template_id})
    tmpl_map = {}
    if tmpl_ids:
        tmpl_result = await db.execute(
            select(PromptTemplate).where(PromptTemplate.id.in_(tmpl_ids))
        )
        tmpl_map = {t.id: t for t in tmpl_result.scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "最新LLM结果"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "LLM结果ID", "检查记录ID", "病历号", "检查日期", "执行时间", "状态", "准确率",
        "ASR结果ID", "ASR模型名称", "ASR转写来源", "LLM模型名称",
        "提示词模板ID", "提示词模板名称", "提示词长度",
        "提示词内容", "ASR原始转写", "实际送入LLM的ASR文本",
        "LLM_右侧卵泡总数", "LLM_右侧卵泡明细", "LLM_左侧卵泡总数", "LLM_左侧卵泡明细",
        "LLM_内膜厚度", "LLM_内膜类型",
        "LLM_右卵巢长", "LLM_右卵巢宽", "LLM_左卵巢长", "LLM_左卵巢宽",
        "LLM_备注", "LLM_总结", "LLM_不确定内容", "LLM_原始返回",
        "真实_右侧卵泡总数", "真实_右侧卵泡明细", "真实_左侧卵泡总数", "真实_左侧卵泡明细",
        "真实_内膜厚度", "真实_内膜类型",
        "真实_右卵巢长", "真实_右卵巢宽", "真实_左卵巢长", "真实_左卵巢宽",
        "真实_备注",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_wrap
        cell.border = thin_border

    def follicles_to_str(follicles):
        if not follicles or not isinstance(follicles, list):
            return "-"
        return "; ".join(f"{f.get('size', '?')}x{f.get('count', '?')}" for f in follicles)

    for row_idx, patient in enumerate(patients, 2):
        llm = latest_llm_by_patient.get(patient.id)
        structured = normalize_structured_result(llm.structured_result or {}) if llm else {}
        right_follicles = structured.get("right_follicles") or []
        left_follicles = structured.get("left_follicles") or []
        asr = asr_map.get(llm.asr_result_id) if llm and llm.asr_result_id else None
        gt = patient.result

        tmpl_id = llm.prompt_template_id if llm else None
        tmpl_name = llm.prompt_template_name if llm else ""
        if not tmpl_name and tmpl_id and tmpl_id in tmpl_map:
            tmpl_name = tmpl_map[tmpl_id].name or ""

        row_data = [
            llm.id if llm else "", patient.id, patient.record_id,
            patient.date_folder.date if patient.date_folder else "",
            llm.created_at.strftime("%Y-%m-%d %H:%M:%S") if llm and llm.created_at else "",
            llm.status if llm else "", llm.accuracy if llm else "",
            llm.asr_result_id if llm else "", asr.asr_model_name if asr else "", "original" if asr else "",
            llm.llm_model_name if llm else "",
            tmpl_id or "", tmpl_name or ("未记录模板名称" if llm else ""),
            len(llm.prompt_content) if llm and llm.prompt_content else 0,
            llm.prompt_content if llm and llm.prompt_content else "",
            asr.full_transcript if asr else "", asr.full_transcript if asr else "",
            structured.get("right_follicle_total", ""), follicles_to_str(right_follicles),
            structured.get("left_follicle_total", ""), follicles_to_str(left_follicles),
            structured.get("endometrium_thickness", ""), structured.get("endometrium_type", ""),
            structured.get("right_ovary_length", ""), structured.get("right_ovary_width", ""),
            structured.get("left_ovary_length", ""), structured.get("left_ovary_width", ""),
            structured.get("remark", ""), llm.summary_text if llm else "",
            structured.get("uncertain_text", ""), llm.raw_output if llm else "",
            gt.right_follicle_total if gt else "", follicles_to_str(gt.right_follicles if gt else []),
            gt.left_follicle_total if gt else "", follicles_to_str(gt.left_follicles if gt else []),
            gt.endometrium_thickness if gt else "", gt.endometrium_type if gt else "",
            gt.right_ovary_length if gt else "", gt.right_ovary_width if gt else "",
            gt.left_ovary_length if gt else "", gt.left_ovary_width if gt else "",
            gt.remark if gt else "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = [
        10, 12, 12, 10, 18, 8, 8,
        10, 15, 12, 15, 12, 15, 10, 50, 50, 50,
        12, 25, 12, 25, 10, 10, 10, 10, 10, 10,
        30, 40, 30, 50,
        12, 25, 12, 25, 10, 10, 10, 10, 10, 10, 30,
    ]
    for idx, w in enumerate(col_widths, 1):
        col_letter = chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"LLM_latest_{len(patients)}_records.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/records")
async def get_records_flat(date: str = None, db: AsyncSession = Depends(get_db)):
    """获取平铺的患者检查记录列表（用于表格展示）"""
    result = await db.execute(
        select(DateFolder)
        .options(
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.segs),
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.result),
        )
        .order_by(DateFolder.date.desc())
    )
    folders = result.unique().scalars().all()

    if date:
        folders = [f for f in folders if f.date == date]

    patient_ids = [p.id for folder in folders for p in folder.patients]
    latest_asr_by_patient = {}
    latest_llm_by_patient = {}
    _llm_asr_cache: dict[int, any] = {}  # patient_id -> PatientAsrResult | None
    _field_review_marks_by_patient: dict[int, list] = {}
    if patient_ids:
        asr_result = await db.execute(
            select(PatientAsrResult)
            .where(PatientAsrResult.patient_id.in_(patient_ids))
            .order_by(
                PatientAsrResult.patient_id.asc(),
                PatientAsrResult.created_at.desc(),
                PatientAsrResult.id.desc(),
            )
        )
        for asr in asr_result.scalars().all():
            if asr.patient_id not in latest_asr_by_patient:
                latest_asr_by_patient[asr.patient_id] = asr

        llm_result = await db.execute(
            select(PatientLlmResult)
            .where(PatientLlmResult.patient_id.in_(patient_ids))
            .order_by(
                PatientLlmResult.patient_id.asc(),
                PatientLlmResult.created_at.desc(),
                PatientLlmResult.id.desc(),
            )
        )
        llm_rows = llm_result.scalars().all()
        # 收集所有 asr_result_id
        llm_asr_ids = list({llm.asr_result_id for llm in llm_rows if llm.asr_result_id})
        # 批量查询关联 ASR
        asr_by_id: dict[int, PatientAsrResult] = {}
        if llm_asr_ids:
            asr_for_llm = await db.execute(
                select(PatientAsrResult).where(PatientAsrResult.id.in_(llm_asr_ids))
            )
            for a in asr_for_llm.scalars().all():
                asr_by_id[a.id] = a
        for llm in llm_rows:
            if llm.patient_id not in latest_llm_by_patient:
                latest_llm_by_patient[llm.patient_id] = llm
                _llm_asr_cache[llm.patient_id] = asr_by_id.get(llm.asr_result_id) if llm.asr_result_id else None

        # 加载字段人工标记
        from app.models import FieldReviewMark
        marks_result = await db.execute(
            select(FieldReviewMark).where(FieldReviewMark.patient_id.in_(patient_ids))
        )
        for m in marks_result.scalars().all():
            _field_review_marks_by_patient.setdefault(m.patient_id, []).append(m)

    def build_ground_truth(result_obj):
        if not result_obj:
            return None
        return {
            "right_follicle_total": result_obj.right_follicle_total,
            "left_follicle_total": result_obj.left_follicle_total,
            "right_follicles": result_obj.right_follicles,
            "left_follicles": result_obj.left_follicles,
            "endometrium_thickness": result_obj.endometrium_thickness,
            "endometrium_type": result_obj.endometrium_type,
            "right_ovary_length": result_obj.right_ovary_length,
            "right_ovary_width": result_obj.right_ovary_width,
            "left_ovary_length": result_obj.left_ovary_length,
            "left_ovary_width": result_obj.left_ovary_width,
            "remark": result_obj.remark,
        }

    def build_latest_asr(asr):
        if not asr:
            return None
        return {
            "id": asr.id,
            "status": asr.status,
            "asr_model_id": asr.asr_model_id,
            "asr_model_name": asr.asr_model_name,
            "provider": asr.provider,
            "created_at": asr.created_at.isoformat() if asr.created_at else None,
            "is_current": asr.is_current,
        }

    def build_latest_llm(llm, result_obj):
        if not llm:
            return None
        structured = normalize_structured_result(llm.structured_result)
        evaluation = llm.evaluation
        evaluation_with_remark = None
        accuracy_without_remark = llm.accuracy
        accuracy_with_remark = None
        gt = build_ground_truth(result_obj)
        if structured and gt:
            evaluation = evaluate_result(structured, gt, include_remark=False)
            evaluation_with_remark = evaluate_result(structured, gt, include_remark=True)
            accuracy_without_remark = evaluation.get("accuracy")
            accuracy_with_remark = evaluation_with_remark.get("accuracy")
        # 关联 ASR 信息
        asr_info = None
        asr = _llm_asr_cache.get(llm.patient_id)
        if asr:
            asr_info = {
                "asr_result_id": asr.id,
                "asr_model_id": asr.asr_model_id,
                "asr_model_name": asr.asr_model_name,
                "asr_provider": asr.provider,
                "asr_status": asr.status,
                "asr_created_at": asr.created_at.isoformat() if asr.created_at else None,
            }
        # 字段人工标记
        marks = _field_review_marks_by_patient.get(llm.patient_id, [])
        return {
            "id": llm.id,
            "status": llm.status,
            "llm_model_name": llm.llm_model_name,
            "prompt_template_name": llm.prompt_template_name,
            "created_at": llm.created_at.isoformat() if llm.created_at else None,
            "accuracy": accuracy_without_remark,
            "accuracy_without_remark": accuracy_without_remark,
            "accuracy_with_remark": accuracy_with_remark,
            "evaluation": evaluation,
            "evaluation_with_remark": evaluation_with_remark,
            "structured_result": structured,
            "ground_truth": gt,
            "asr_result_id": asr_info["asr_result_id"] if asr_info else llm.asr_result_id,
            "asr_model_id": asr_info["asr_model_id"] if asr_info else None,
            "asr_model_name": asr_info["asr_model_name"] if asr_info else None,
            "asr_provider": asr_info["asr_provider"] if asr_info else None,
            "asr_status": asr_info["asr_status"] if asr_info else None,
            "asr_created_at": asr_info["asr_created_at"] if asr_info else None,
            "field_review_marks": [
                {
                    "id": m.id,
                    "patient_id": m.patient_id,
                    "field_group": m.field_group,
                    "field_key": m.field_key,
                    "mark_type": m.mark_type,
                    "reason": m.reason,
                    "note": m.note,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                    "updated_at": m.updated_at.isoformat() if m.updated_at else None,
                }
                for m in marks
            ],
        }

    records = []
    for folder in folders:
        for p in folder.patients:
            records.append({
                "id": p.id,
                "record_id": p.record_id,
                "date": folder.date,
                "date_folder_id": folder.id,
                "timestamp_folder": p.timestamp_folder,
                "note": p.note or "",
                "has_audio": len(p.segs) > 0,
                "has_result": p.result is not None,
                "segs": [
                    {
                        "id": s.id,
                        "seg_index": s.seg_index,
                        "filename": s.filename,
                        "duration": s.duration,
                        "file_path": s.file_path,
                        "file_size": s.file_size,
                    }
                    for s in sorted(p.segs, key=lambda x: x.seg_index)
                ],
                "result": None,
                "latest_asr": build_latest_asr(latest_asr_by_patient.get(p.id)),
                "latest_llm": build_latest_llm(latest_llm_by_patient.get(p.id), p.result),
            })
            if p.result:
                records[-1]["result"] = {
                    "id": p.result.id,
                    "right_follicles": p.result.right_follicles,
                    "left_follicles": p.result.left_follicles,
                    "right_follicle_total": p.result.right_follicle_total,
                    "left_follicle_total": p.result.left_follicle_total,
                    "endometrium_thickness": p.result.endometrium_thickness,
                    "endometrium_type": p.result.endometrium_type,
                    "right_ovary_length": p.result.right_ovary_length,
                    "right_ovary_width": p.result.right_ovary_width,
                    "left_ovary_length": p.result.left_ovary_length,
                    "left_ovary_width": p.result.left_ovary_width,
                    "remark": p.result.remark,
                }

    return records


@router.put("/patient/{patient_id}/note")
async def update_patient_note(
    patient_id: int,
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
):
    """更新检查记录人工备注/标注，不影响真实 B 超备注。"""
    patient = await db.get(PatientRecord, patient_id)
    if not patient:
        raise HTTPException(status_code=404, detail="检查记录不存在")

    patient.note = str(data.get("note") or "").strip() or None
    await db.commit()
    await db.refresh(patient)
    return {"id": patient.id, "note": patient.note or ""}


@router.delete("/patient/{patient_id}")
async def delete_patient(patient_id: int, db: AsyncSession = Depends(get_db)):
    """删除患者记录（包括关联的 B超结果）"""
    import asyncio

    def _do_delete():
        import sqlite3
        db_path = settings.DATABASE_URL.replace("sqlite:///", "")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        # Delete associated result
        cursor.execute("DELETE FROM b_ultra_results WHERE patient_id = ?", (patient_id,))
        # Delete associated segs
        cursor.execute("DELETE FROM audio_segs WHERE patient_id = ?", (patient_id,))
        # Delete patient
        cursor.execute("DELETE FROM patient_records WHERE id = ?", (patient_id,))
        conn.commit()
        deleted = cursor.rowcount
        conn.close()
        return deleted

    deleted = await asyncio.to_thread(_do_delete)
    if deleted == 0:
        raise HTTPException(status_code=404, detail="患者不存在")
    return {"message": "已删除", "id": patient_id}


@router.get("/patients")
async def get_patients_list(date: str = None, db: AsyncSession = Depends(get_db)):
    """获取患者检查列表（按病历号分组，含录音和结果），支持按日期筛选"""
    result = await db.execute(
        select(DateFolder)
        .options(
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.segs),
            selectinload(DateFolder.patients)
            .selectinload(PatientRecord.result),
        )
        .order_by(DateFolder.date.desc())
    )
    folders = result.unique().scalars().all()

    # Filter by date if provided
    if date:
        folders = [f for f in folders if f.date == date]

    # Group by record_id
    patients_map: dict[str, list] = {}
    for folder in folders:
        for p in folder.patients:
            record = {
                "id": p.id,
                "record_id": p.record_id,
                "date": folder.date,
                "date_folder_id": folder.id,
                "timestamp_folder": p.timestamp_folder,
                "note": p.note or "",
                "segs": [
                    {
                        "id": s.id,
                        "seg_index": s.seg_index,
                        "filename": s.filename,
                        "duration": s.duration,
                        "file_path": s.file_path,
                        "file_size": s.file_size,
                    }
                    for s in sorted(p.segs, key=lambda x: x.seg_index)
                ],
                "result": None,
            }
            if p.result:
                record["result"] = {
                    "id": p.result.id,
                    "right_follicles": p.result.right_follicles,
                    "left_follicles": p.result.left_follicles,
                    "right_follicle_total": p.result.right_follicle_total,
                    "left_follicle_total": p.result.left_follicle_total,
                    "endometrium_thickness": p.result.endometrium_thickness,
                    "endometrium_type": p.result.endometrium_type,
                    "right_ovary_length": p.result.right_ovary_length,
                    "right_ovary_width": p.result.right_ovary_width,
                    "left_ovary_length": p.result.left_ovary_length,
                    "left_ovary_width": p.result.left_ovary_width,
                    "remark": p.result.remark,
                }
            if p.record_id not in patients_map:
                patients_map[p.record_id] = []
            patients_map[p.record_id].append(record)

    # Output sorted by date
    output = []
    for record_id, records in patients_map.items():
        output.append({
            "record_id": record_id,
            "examinations": sorted(records, key=lambda x: x["date"], reverse=True),
        })
    return output


def self_find_audio_dir(record_path: str) -> str | None:
    """查找病历号目录下的 audio 文件夹"""
    for root, dirs, files in os.walk(record_path):
        if os.path.basename(root) == "audio":
            return root
    return None


def self_parse_seg_index(filename: str) -> int:
    """从 seg-0001.wav 解析索引"""
    import re
    m = re.search(r"seg-(\d+)", filename)
    return int(m.group(1)) if m else 0
