"""
实验控制平面 API 路由
"""
import json
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from typing import Optional

from app.database import get_db
from app.models import ModelConfig, DateFolder, PatientRecord
from app.models.experiment import (
    ExperimentBatch, ExperimentCombination, ExperimentTask,
    BatchStatus, TaskStatus, TaskStage,
)
from app.schemas import (
    ExperimentBatchCreate, ExperimentBatchOut, ExperimentDetailOut,
    ExperimentListResponse, ExperimentCombinationCreate,
    ExperimentCombinationUpdate, ExperimentCombinationOut,
    ExperimentPatientScopeUpdate, ExperimentControlAction,
    ExperimentTaskSummary, ExperimentMetrics,
)
from app.services.experiment_planner import plan_tasks, invalidate_tasks
from app.services.experiment_metrics import calculate_metrics
from app.services.experiment_metrics import get_batch_metrics

router = APIRouter()


@router.get("")
async def list_experiments(db: AsyncSession = Depends(get_db)):
    """获取实验批次列表 (含聚合指标)"""
    result = await db.execute(
        select(ExperimentBatch)
        .options(
            selectinload(ExperimentBatch.combinations).selectinload(ExperimentCombination.asr_model),
            selectinload(ExperimentBatch.combinations).selectinload(ExperimentCombination.llm_model),
            selectinload(ExperimentBatch.tasks),
        )
        .order_by(ExperimentBatch.created_at.desc())
    )
    batches = result.scalars().all()

    output = []
    for b in batches:
        # JSON 字段处理
        pids = b.selected_patient_ids or []
        if isinstance(pids, str):
            try: pids = json.loads(pids)
            except: pids = []
        dates = b.selected_dates or []
        if isinstance(dates, str):
            try: dates = json.loads(dates)
            except: dates = []

        # 聚合指标
        tasks = b.tasks or []
        m = calculate_metrics(tasks)

        # 单一组合信息（新设计）
        first_combo = (b.combinations or [None])[0]
        asr_model_name = first_combo.asr_model.name if first_combo and first_combo.asr_model else ""
        llm_model_name = first_combo.llm_model.name if first_combo and first_combo.llm_model else ""
        prompt_template_name = first_combo.prompt_name if first_combo else ""

        # ASR 来源统计
        asr_reuse = sum(1 for t in tasks if t.asr_source == "reused")
        asr_generated = sum(1 for t in tasks if t.asr_source == "generated")
        asr_failed = sum(1 for t in tasks if t.asr_source == "failed")

        output.append({
            "id": b.id,
            "name": b.name,
            "status": b.status,
            "remark": b.remark or "",
            "selected_dates": dates,
            "selected_patient_ids": pids,
            "total_tasks": b.total_tasks or 0,
            "success_count": b.success_count or 0,
            "failure_count": b.failure_count or 0,
            "created_at": b.created_at.isoformat() if b.created_at else None,
            "updated_at": b.updated_at.isoformat() if b.updated_at else None,
            "started_at": b.started_at.isoformat() if b.started_at else None,
            "completed_at": b.completed_at.isoformat() if b.completed_at else None,
            "patient_count": len(pids) if isinstance(pids, list) else 0,
            "combination_count": len(b.combinations or []),
            "metrics": m,
            "field_accuracy": m.get("field_accuracy", {}),
            "asr_model_name": asr_model_name,
            "llm_model_name": llm_model_name,
            "prompt_template_name": prompt_template_name,
            "asr_reuse_count": asr_reuse,
            "asr_generated_count": asr_generated,
            "asr_failed_count": asr_failed,
            "asr_reuse_rate": asr_reuse / max(len(tasks), 1),
        })
    return output


@router.post("", response_model=ExperimentBatchOut)
async def create_experiment(
    data: ExperimentBatchCreate,
    db: AsyncSession = Depends(get_db),
):
    """创建实验批次（新设计：单组合，立即创建 pending tasks）"""
    # 校验：如果配置了 LLM，必须有提示词
    if data.llm_model_id and not data.prompt_template and not data.prompt_template_id:
        raise HTTPException(status_code=400, detail="使用 LLM 时必须配置提示词模板或提示词内容")

    batch = ExperimentBatch(
        name=data.name,
        description=data.description,
        remark=data.remark,
        selected_dates=data.selected_dates,
        selected_patient_ids=data.selected_patient_ids,
        status=BatchStatus.PENDING.value,
    )
    db.add(batch)
    await db.flush()  # 获取 batch.id

    # 解析提示词（优先使用传入内容，否则从模板查询）
    prompt_name = data.prompt_name or ""
    prompt_template = data.prompt_template or ""

    if data.prompt_template_id and not prompt_template:
        from app.models import PromptTemplate
        tmpl = await db.get(PromptTemplate, data.prompt_template_id)
        if tmpl:
            prompt_name = tmpl.name
            prompt_template = tmpl.content

    # 最终校验：如果配置了 LLM，提示词内容必须非空
    if data.llm_model_id and not prompt_template:
        raise HTTPException(
            status_code=400,
            detail="使用 LLM 时必须配置有效的提示词模板或提示词内容"
        )

    # 创建单条 combination 记录
    if data.asr_model_id:
        combo = ExperimentCombination(
            batch_id=batch.id,
            asr_model_id=data.asr_model_id,
            llm_model_id=data.llm_model_id,
            prompt_name=prompt_name,
            prompt_template=prompt_template,
            hotwords=data.hotwords or [],
            enabled=True,
        )
        db.add(combo)
        await db.flush()

        # 立即创建 pending tasks（根据 selected_dates + selected_patient_ids）
        if batch.selected_dates and batch.selected_patient_ids:
            patient_records = await db.execute(
                select(PatientRecord)
                .join(DateFolder)
                .where(
                    DateFolder.date.in_(batch.selected_dates),
                    PatientRecord.record_id.in_(batch.selected_patient_ids),
                )
            )
            patients = patient_records.scalars().all()
            for p in patients:
                task = ExperimentTask(
                    batch_id=batch.id,
                    combination_id=combo.id,
                    patient_id=p.id,
                    stage=TaskStage.ASR.value,
                    status=TaskStatus.PENDING.value,
                )
                db.add(task)
            batch.total_tasks = len(patients)

    await db.commit()
    await db.refresh(batch)
    return batch


@router.delete("/{batch_id}")
async def delete_experiment(batch_id: int, db: AsyncSession = Depends(get_db)):
    """删除实验批次"""
    batch = await db.get(ExperimentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")
    await db.delete(batch)
    await db.commit()
    return {"message": "已删除", "id": batch_id}


@router.get("/{batch_id}", response_model=ExperimentDetailOut)
async def get_experiment(
    batch_id: int,
    db: AsyncSession = Depends(get_db),
):
    """获取实验详情"""
    result = await db.execute(
        select(ExperimentBatch)
        .options(selectinload(ExperimentBatch.combinations))
        .where(ExperimentBatch.id == batch_id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")
    # 手动构建响应以确保 JSON 字段正确序列化
    first_combo = batch.combinations[0] if batch.combinations else None
    return {
        "id": batch.id,
        "name": batch.name,
        "description": batch.description,
        "remark": batch.remark or "",
        "selected_dates": batch.selected_dates or [],
        "selected_patient_ids": batch.selected_patient_ids or [],
        "status": batch.status,
        "total_tasks": batch.total_tasks,
        "success_count": batch.success_count,
        "failure_count": batch.failure_count,
        "created_at": batch.created_at,
        "updated_at": batch.updated_at,
        "started_at": batch.started_at,
        "completed_at": batch.completed_at,
        "combinations": [
            {
                "id": c.id,
                "batch_id": c.batch_id,
                "asr_model_id": c.asr_model_id,
                "llm_model_id": c.llm_model_id,
                "prompt_name": c.prompt_name or "",
                "prompt_template": c.prompt_template or "",
                "hotwords": c.hotwords or [],
                "enabled": c.enabled,
                "created_at": c.created_at,
            }
            for c in batch.combinations
        ],
        # 单一组合便捷字段
        "asr_model_id": first_combo.asr_model_id if first_combo else None,
        "llm_model_id": first_combo.llm_model_id if first_combo else None,
        "prompt_name": first_combo.prompt_name if first_combo else "",
        "prompt_template_name": first_combo.prompt_name if first_combo else "",
        "hotwords": first_combo.hotwords if first_combo else [],
    }


@router.post("/{batch_id}/combinations")
async def add_combination(
    batch_id: int,
    data: ExperimentCombinationCreate,
    db: AsyncSession = Depends(get_db),
):
    """添加实验组合（已禁用：单组合设计）"""
    # 检查是否已有组合
    existing = await db.execute(
        select(ExperimentCombination).where(ExperimentCombination.batch_id == batch_id)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail="当前版本一个实验只允许一个组合，如需比较多个组合请新建实验。"
        )
    raise HTTPException(status_code=400, detail="当前版本不支持添加组合，请在创建实验时配置。")


@router.put("/{batch_id}/combinations/{combo_id}", response_model=ExperimentCombinationOut)
async def update_combination(
    batch_id: int,
    combo_id: int,
    data: ExperimentCombinationUpdate,
    db: AsyncSession = Depends(get_db),
):
    """更新实验组合"""
    result = await db.execute(
        select(ExperimentCombination).where(
            ExperimentCombination.id == combo_id,
            ExperimentCombination.batch_id == batch_id,
        )
    )
    combo = result.scalar_one_or_none()
    if not combo:
        raise HTTPException(status_code=404, detail="组合不存在")

    # Store old values for invalidation
    old_values = {
        "asr_model_id": combo.asr_model_id,
        "llm_model_id": combo.llm_model_id,
        "prompt_template": combo.prompt_template,
        "hotwords": combo.hotwords,
    }

    # Update fields
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(combo, field, value)

    await db.commit()

    # Invalidate affected tasks
    from app.services.experiment_planner import changed_stage
    new_values = {
        "asr_model_id": combo.asr_model_id,
        "llm_model_id": combo.llm_model_id,
        "prompt_template": combo.prompt_template,
        "hotwords": combo.hotwords,
    }
    stage = changed_stage(old_values, new_values)
    if stage:
        await invalidate_tasks(db, combo_id, stage)

    await db.refresh(combo)
    return combo


@router.delete("/{batch_id}/combinations/{combo_id}")
async def delete_combination(
    batch_id: int,
    combo_id: int,
    db: AsyncSession = Depends(get_db),
):
    """删除实验组合"""
    from sqlalchemy import delete as sa_delete
    result = await db.execute(
        select(ExperimentCombination).where(
            ExperimentCombination.id == combo_id,
            ExperimentCombination.batch_id == batch_id,
        )
    )
    combo = result.scalar_one_or_none()
    if not combo:
        raise HTTPException(status_code=404, detail="组合不存在")

    # Delete associated tasks
    await db.execute(
        sa_delete(ExperimentTask).where(ExperimentTask.combination_id == combo_id)
    )
    # Delete combination
    await db.delete(combo)
    await db.commit()
    return {"message": "已删除", "id": combo_id}


@router.put("/{batch_id}/patients", response_model=ExperimentBatchOut)
async def update_patient_scope(
    batch_id: int,
    data: ExperimentPatientScopeUpdate,
    db: AsyncSession = Depends(get_db),
):
    """更新患者范围"""
    batch = await db.get(ExperimentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")

    batch.selected_patient_ids = data.selected_patient_ids
    await db.commit()
    await db.refresh(batch)
    return batch


@router.post("/{batch_id}/start", response_model=dict)
async def start_experiment(
    batch_id: int,
    data: ExperimentControlAction = ExperimentControlAction(),
    db: AsyncSession = Depends(get_db),
):
    """启动实验（单组合，快速返回）"""
    result = await db.execute(
        select(ExperimentBatch)
        .where(ExperimentBatch.id == batch_id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")

    # 检查是否已有组合
    combo_result = await db.execute(
        select(ExperimentCombination).where(
            ExperimentCombination.batch_id == batch_id,
            ExperimentCombination.enabled == True,
        )
    )
    combo = combo_result.scalar_one_or_none()
    if not combo:
        raise HTTPException(status_code=400, detail="没有启用的实验组合，请先配置 ASR/LLM/提示词")

    # 检查任务状态
    task_stats = await db.execute(
        select(ExperimentTask.status, func.count(ExperimentTask.id))
        .where(ExperimentTask.batch_id == batch_id)
        .group_by(ExperimentTask.status)
    )
    counts = {row[0]: row[1] for row in task_stats.all()}
    total = sum(counts.values())
    pending = counts.get("pending", 0)
    running = counts.get("running", 0)
    success = counts.get("success", 0)
    failed = counts.get("failed", 0)

    # 如果没有任务，尝试创建
    if total == 0 and batch.selected_dates and batch.selected_patient_ids:
        patient_records = await db.execute(
            select(PatientRecord)
            .join(DateFolder)
            .where(
                DateFolder.date.in_(batch.selected_dates),
                PatientRecord.record_id.in_(batch.selected_patient_ids),
            )
        )
        patients = patient_records.scalars().all()
        for p in patients:
            task = ExperimentTask(
                batch_id=batch.id,
                combination_id=combo.id,
                patient_id=p.id,
                stage=TaskStage.ASR.value,
                status=TaskStatus.PENDING.value,
            )
            db.add(task)
        pending = len(patients)
        batch.total_tasks = pending

    # 如果所有任务已完成，直接设置终态
    if pending == 0 and running == 0:
        if failed == 0 and success > 0:
            batch.status = BatchStatus.COMPLETED.value
        elif failed > 0:
            batch.status = BatchStatus.PARTIAL.value
        else:
            batch.status = BatchStatus.PENDING.value
        await db.commit()
        return {"batch_id": batch_id, "total_tasks": batch.total_tasks or total, "status": batch.status}

    batch.status = BatchStatus.RUNNING.value
    await db.commit()

    return {"batch_id": batch_id, "total_tasks": batch.total_tasks or total, "status": "running"}


@router.get("/{batch_id}/progress")
async def get_progress(batch_id: int, db: AsyncSession = Depends(get_db)):
    """获取实验实时进度"""
    from sqlalchemy import func
    batch = await db.get(ExperimentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")

    # Count by status
    result = await db.execute(
        select(ExperimentTask.status, func.count(ExperimentTask.id))
        .where(ExperimentTask.batch_id == batch_id)
        .group_by(ExperimentTask.status)
    )
    counts = {row[0]: row[1] for row in result.all()}

    total = batch.total_tasks or 1
    success = counts.get("success", 0)
    failed = counts.get("failed", 0)
    running = counts.get("running", 0)
    pending = counts.get("pending", 0)

    return {
        "batch_id": batch_id,
        "total": total,
        "success": success,
        "failed": failed,
        "running": running,
        "pending": pending,
        "percent": round((success + failed) / total * 100, 1) if total > 0 else 0,
        "status": batch.status,
    }


@router.post("/{batch_id}/pause", response_model=ExperimentBatchOut)
async def pause_experiment(
    batch_id: int,
    data: ExperimentControlAction = ExperimentControlAction(),
    db: AsyncSession = Depends(get_db),
):
    """暂停实验"""
    batch = await db.get(ExperimentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")
    if batch.status != BatchStatus.RUNNING.value:
        raise HTTPException(status_code=400, detail="实验未在运行")

    batch.status = BatchStatus.PAUSED.value
    await db.commit()
    await db.refresh(batch)
    return batch


@router.post("/{batch_id}/resume", response_model=ExperimentBatchOut)
async def resume_experiment(
    batch_id: int,
    data: ExperimentControlAction = ExperimentControlAction(),
    db: AsyncSession = Depends(get_db),
):
    """继续实验"""
    batch = await db.get(ExperimentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")
    if batch.status != BatchStatus.PAUSED.value:
        raise HTTPException(status_code=400, detail="实验未暂停")

    batch.status = BatchStatus.RUNNING.value
    await db.commit()
    await db.refresh(batch)
    return batch


@router.post("/{batch_id}/cancel", response_model=ExperimentBatchOut)
async def cancel_experiment(
    batch_id: int,
    data: ExperimentControlAction = ExperimentControlAction(),
    db: AsyncSession = Depends(get_db),
):
    """取消实验"""
    batch = await db.get(ExperimentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")

    batch.status = BatchStatus.CANCELLED.value

    # Cancel pending/running tasks
    result = await db.execute(
        select(ExperimentTask).where(
            ExperimentTask.batch_id == batch_id,
            ExperimentTask.status.in_([TaskStatus.PENDING.value, TaskStatus.RUNNING.value]),
        )
    )
    for task in result.scalars().all():
        task.status = TaskStatus.CANCELLED.value

    await db.commit()
    await db.refresh(batch)
    return batch


@router.post("/{batch_id}/retry", response_model=dict)
async def retry_failures(
    batch_id: int,
    data: ExperimentControlAction = ExperimentControlAction(),
    db: AsyncSession = Depends(get_db),
):
    """重试失败任务"""
    result = await db.execute(
        select(ExperimentTask).where(
            ExperimentTask.batch_id == batch_id,
            ExperimentTask.status == TaskStatus.FAILED.value,
        )
    )
    failed_tasks = result.scalars().all()

    for task in failed_tasks:
        task.status = TaskStatus.PENDING.value
        task.retry_count = 0
        task.error_type = None
        task.error_message = None

    await db.commit()
    return {"batch_id": batch_id, "retried": len(failed_tasks)}


@router.get("/{batch_id}/tasks")
async def list_tasks(
    batch_id: int,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """获取实验任务列表（含患者信息 + ground_truth）"""
    from app.services.parser import normalize_structured_result
    query = (
        select(ExperimentTask)
        .options(
            selectinload(ExperimentTask.patient).selectinload(PatientRecord.date_folder),
            selectinload(ExperimentTask.patient).selectinload(PatientRecord.result),
            selectinload(ExperimentTask.combination).selectinload(ExperimentCombination.asr_model),
            selectinload(ExperimentTask.combination).selectinload(ExperimentCombination.llm_model),
        )
        .where(ExperimentTask.batch_id == batch_id)
    )
    if status:
        query = query.where(ExperimentTask.status == status)
    result = await db.execute(query.order_by(ExperimentTask.id))
    tasks = result.scalars().all()

    output = []
    for t in tasks:
        # 构建 ground_truth (可能为 null)
        ground_truth = None
        if t.patient and t.patient.result:
            r = t.patient.result
            ground_truth = {
                "right_follicles": r.right_follicles,
                "left_follicles": r.left_follicles,
                "right_follicle_total": r.right_follicle_total,
                "left_follicle_total": r.left_follicle_total,
                "endometrium_thickness": r.endometrium_thickness,
                "endometrium_type": r.endometrium_type,
                "right_ovary_length": r.right_ovary_length,
                "right_ovary_width": r.right_ovary_width,
                "left_ovary_length": r.left_ovary_length,
                "left_ovary_width": r.left_ovary_width,
                "remark": r.remark,
            }

        output.append({
            "id": t.id,
            "batch_id": t.batch_id,
            "combination_id": t.combination_id,
            "exam_record_id": t.patient_id,  # patient_id 实际是 exam_record_id
            "patient_id": t.patient_id,      # 保留兼容
            "record_id": t.patient.record_id if t.patient else None,
            "date": t.patient.date_folder.date if t.patient and t.patient.date_folder else None,
            "stage": t.stage,
            "status": t.status,
            "retry_count": t.retry_count,
            "accuracy": t.accuracy,
            "total_duration": t.total_duration,
            "error_type": t.error_type,
            "error_message": t.error_message,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            # ASR 快照字段
            "asr_result_id": t.asr_result_id,
            "asr_source": t.asr_source,
            "asr_model_name": t.asr_model_name,
            "asr_results": t.asr_results,
            "full_transcript": t.full_transcript,
            # LLM 快照字段
            "llm_result_id": t.llm_result_id,
            "llm_model_name": t.llm_model_name,
            "prompt_template_name": t.prompt_template_name,
            "llm_raw_output": t.llm_raw_output,
            "structured_result": normalize_structured_result(t.structured_result),
            "summary_text": t.summary_text,
            # 评估
            "evaluation": t.evaluation,
            "ground_truth": ground_truth,
            # 组合信息
            "combination_asr_name": t.combination.asr_model.name if t.combination and t.combination.asr_model else "",
            "combination_llm_name": t.combination.llm_model.name if t.combination and t.combination.llm_model else "",
            "combination_prompt_name": t.combination.prompt_name or "",
        })
    return output


@router.get("/{batch_id}/metrics")
async def get_metrics(batch_id: int, db: AsyncSession = Depends(get_db)):
    """获取实验指标"""
    return await get_batch_metrics(db, batch_id)


@router.get("/{batch_id}/results")
async def get_results(batch_id: int, db: AsyncSession = Depends(get_db)):
    """获取实验结果列表"""
    result = await db.execute(
        select(ExperimentTask).where(ExperimentTask.batch_id == batch_id)
    )
    return result.scalars().all()


@router.get("/{batch_id}/export")
async def export_experiment(batch_id: int, db: AsyncSession = Depends(get_db)):
    """导出实验全部任务结果为 Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response
    from app.models import BUltraResult, PromptTemplate, PatientAsrResult, PatientLlmResult
    from app.services.parser import normalize_structured_result

    batch = await db.get(ExperimentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="实验不存在")

    # 获取组合信息
    combo = await db.execute(
        select(ExperimentCombination).where(ExperimentCombination.batch_id == batch_id)
    )
    combination = combo.scalar_one_or_none()

    # 查询任务
    tasks_result = await db.execute(
        select(ExperimentTask)
        .options(
            selectinload(ExperimentTask.patient).selectinload(PatientRecord.date_folder),
            selectinload(ExperimentTask.patient).selectinload(PatientRecord.result),
            selectinload(ExperimentTask.combination).selectinload(ExperimentCombination.asr_model),
            selectinload(ExperimentTask.combination).selectinload(ExperimentCombination.llm_model),
        )
        .where(ExperimentTask.batch_id == batch_id)
        .order_by(ExperimentTask.id)
    )
    tasks = tasks_result.scalars().all()
    if not tasks:
        raise HTTPException(status_code=404, detail="实验暂无任务")

    # 准备辅助数据
    patient_ids = [t.patient_id for t in tasks if t.patient_id]
    gt_results = await db.execute(
        select(BUltraResult).where(BUltraResult.patient_id.in_(patient_ids))
    )
    gt_map = {g.patient_id: g for g in gt_results.scalars().all()}

    asr_ids = [t.asr_result_id for t in tasks if t.asr_result_id]
    asr_results = await db.execute(
        select(PatientAsrResult).where(PatientAsrResult.id.in_(asr_ids))
    )
    asr_map = {a.id: a for a in asr_results.scalars().all()}

    llm_ids = [t.llm_result_id for t in tasks if t.llm_result_id]
    llm_results = await db.execute(
        select(PatientLlmResult).where(PatientLlmResult.id.in_(llm_ids))
    )
    llm_map = {l.id: l for l in llm_results.scalars().all()}

    # 导出
    wb = Workbook()
    ws = wb.active
    ws.title = "实验结果"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "实验ID", "实验名称", "组合ID", "任务ID", "病历号", "检查日期", "状态", "准确率",
        "错误信息", "执行耗时(s)",
        "ASR来源", "ASR模型名称", "ASR结果ID", "LLM模型名称",
        "提示词模板名称", "提示词长度", "提示词内容", "ASR转写文本",
        "LLM_右侧卵泡总数", "LLM_右侧卵泡明细", "LLM_左侧卵泡总数", "LLM_左侧卵泡明细",
        "LLM_内膜厚度", "LLM_内膜类型",
        "LLM_右卵巢长", "LLM_右卵巢宽", "LLM_左卵巢长", "LLM_左卵巢宽",
        "LLM_备注", "LLM_总结",
        "真实_右侧卵泡总数", "真实_右侧卵泡明细", "真实_左侧卵泡总数", "真实_左侧卵泡明细",
        "真实_内膜厚度", "真实_内膜类型",
        "真实_右卵巢长", "真实_右卵巢宽", "真实_左卵巢长", "真实_左卵巢宽", "真实_备注",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_wrap
        cell.border = thin_border

    def follicles_to_str(follicles: any) -> str:
        if not follicles or not isinstance(follicles, list):
            return "-"
        return "; ".join(f"{f.get('size', '?')}x{f.get('count', '?')}" for f in follicles)

    # 获取提示词模板信息
    prompt_content = ""
    prompt_name = ""
    if combination:
        prompt_name = combination.prompt_name or ""
        prompt_content = combination.prompt_template or ""

    batch_name = batch.name or f"实验{batch_id}"

    for row_idx, t in enumerate(tasks, 2):
        patient = t.patient
        record_id = patient.record_id if patient else f"record_{t.patient_id}"
        date_str = patient.date_folder.date if patient and patient.date_folder else ""

        # ASR
        asr_record = asr_map.get(t.asr_result_id) if t.asr_result_id else None
        asr_model_name = ""
        asr_transcript = ""
        if asr_record:
            asr_model_name = asr_record.asr_model_name or t.asr_model_name or ""
            asr_transcript = asr_record.full_transcript or t.full_transcript or ""
        else:
            asr_model_name = t.asr_model_name or ""
            asr_transcript = t.full_transcript or ""
        asr_source = t.asr_source or "-"

        # LLM
        llm_record = llm_map.get(t.llm_result_id) if t.llm_result_id else None
        structured = normalize_structured_result(t.structured_result or (llm_record.structured_result if llm_record else None) or {})
        llm_model_name = t.llm_model_name or (llm_record.llm_model_name if llm_record else "") or ""
        llm_summary = t.summary_text or (llm_record.summary_text if llm_record else "") or ""
        llm_raw = t.llm_raw_output or (llm_record.raw_output if llm_record else "") or ""

        # Ground truth
        gt = gt_map.get(t.patient_id)

        row_data = [
            batch_id, batch_name, t.combination_id or "", t.id,
            record_id, date_str, t.status or "pending",
            t.accuracy if t.accuracy is not None else "",
            t.error_message or "",
            t.total_duration or "",
            asr_source, asr_model_name, t.asr_result_id or "",
            llm_model_name,
            prompt_name, len(prompt_content) if prompt_content else 0, prompt_content,
            asr_transcript,
            structured.get("right_follicle_total", ""),
            follicles_to_str(structured.get("right_follicles") if isinstance(structured, dict) else None),
            structured.get("left_follicle_total", ""),
            follicles_to_str(structured.get("left_follicles") if isinstance(structured, dict) else None),
            structured.get("endometrium_thickness", ""),
            structured.get("endometrium_type", ""),
            structured.get("right_ovary_length", ""),
            structured.get("right_ovary_width", ""),
            structured.get("left_ovary_length", ""),
            structured.get("left_ovary_width", ""),
            structured.get("remark", ""),
            llm_summary,
            gt.right_follicle_total if gt else "",
            follicles_to_str(gt.right_follicles if gt else None),
            gt.left_follicle_total if gt else "",
            follicles_to_str(gt.left_follicles if gt else None),
            gt.endometrium_thickness if gt else "",
            gt.endometrium_type if gt else "",
            gt.right_ovary_length if gt else "",
            gt.right_ovary_width if gt else "",
            gt.left_ovary_length if gt else "",
            gt.left_ovary_width if gt else "",
            gt.remark if gt else "",
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = [
        8, 15, 8, 8, 12, 10, 10, 8, 30, 10,
        10, 15, 10, 15,
        15, 10, 50, 50,
        12, 25, 12, 25, 10, 10, 10, 10, 10, 10,
        30, 40,
        12, 25, 12, 25, 10, 10, 10, 10, 10, 10, 30,
    ]
    for idx, w in enumerate(col_widths, 1):
        if idx <= 26:
            col_letter = chr(64 + idx)
        elif idx <= 52:
            col_letter = "A" + chr(64 + idx - 26)
        else:
            col_letter = "B" + chr(64 + idx - 52)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = batch_name.encode("ascii", "ignore").decode() or f"batch_{batch_id}"
    filename = f"Experiment_{batch_id}_{safe_name}_export.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
