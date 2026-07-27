"""
检查记录级 ASR/LLM 持久化结果路由

业务语义说明:
- 本模块所有 patient_id 实际为 exam_record_id (= patient_records.id)
- record_id (病历号) 可跨日期重复, 不能作为结果关联键
- 每个检查记录 (exam_record) 有独立的 ASR/LLM 结果

GET  /api/patients/{patient_id}/asr/stream    SSE 流式 ASR, 保存到 patient_asr_results
GET  /api/patients/{patient_id}/asr-results
GET  /api/patients/{patient_id}/asr-current
PUT  /api/patients/{patient_id}/asr-results/{result_id}/current

POST /api/patients/{patient_id}/llm/run      LLM 结构化提取, 保存到 patient_llm_results
GET  /api/patients/{patient_id}/llm-results
GET  /api/patients/{patient_id}/llm-current
PUT  /api/patients/{patient_id}/llm-results/{result_id}/current
"""
import asyncio
import json
import io
import time
import re
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy import select, update, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from loguru import logger

from app.config import resolve_hotwords
from app.database import get_db, AsyncSessionLocal
from app.models import (
    PatientRecord, AudioSeg, ModelConfig,
    PatientAsrResult, PatientLlmResult, AsrReferenceTranscript,
)
from app.services.asr import create_asr
from app.services.asr_input import build_asr_audio_inputs
from app.services.test_executor import TestExecutor

router = APIRouter()

ASR_TRANSCRIBE_TIMEOUT_SECONDS = 600
ASR_STALE_RUNNING_SECONDS = 30 * 60
_RUNNING_ASR_TASKS: set[asyncio.Task] = set()
_NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")


ASR_PARAM_OVERRIDE_ALLOWLIST = {
    "audio_input_mode",
    "recognition_mode",
    "endpoint_mode",
    "language",
    "stream",
    "merge_group_size",
    "max_base64_mb",
    "result_type",
    "enable_itn",
    "enable_punc",
    "enable_ddc",
    "show_utterances",
    "enable_nonstream",
    "enable_speaker_info",
    "end_window_size",
    "vad_segment_duration",
    "force_to_speech_time",
    "enable_auto_lang",
    "hotwords",
    "context_text",
    "use_context_hotwords",
    "context_mode",
    "use_boosting_table",
    "use_correct_table",
    "boosting_table_id",
    "boosting_table_name",
    "correct_table_id",
    "correct_table_name",
    "frame_size",
    "send_interval",
    "resource_id",
    "task_timeout_seconds",
    "receive_timeout_seconds",
    "total_timeout_seconds",
}


def _parse_asr_params_override(raw: Optional[str]) -> dict:
    """解析 ASR 优化评估传入的临时参数，只允许覆盖非密钥类参数。"""
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="params_override 不是合法 JSON") from exc
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=400, detail="params_override 必须是对象")

    cleaned: dict = {}
    for key, value in parsed.items():
        if key not in ASR_PARAM_OVERRIDE_ALLOWLIST:
            continue
        if value is None or value == "":
            continue
        if key == "hotwords":
            if isinstance(value, str):
                words = [item.strip() for item in value.replace("，", ",").split(",") if item.strip()]
                if words:
                    cleaned[key] = words
            elif isinstance(value, list):
                words = [str(item).strip() for item in value if str(item).strip()]
                if words:
                    cleaned[key] = words
            continue
        cleaned[key] = value
    return cleaned


def _apply_asr_feature_switches(params: dict) -> dict:
    """按启用开关决定是否把平台词表/上下文热词传给 ASR 服务。"""
    next_params = dict(params or {})
    if next_params.get("use_boosting_table", True) is False:
        next_params.pop("boosting_table_id", None)
        next_params.pop("boosting_table_name", None)
    if next_params.get("use_correct_table", True) is False:
        next_params.pop("correct_table_id", None)
        next_params.pop("correct_table_name", None)
    if next_params.get("use_context_hotwords", True) is False:
        next_params.pop("hotwords", None)
        next_params.pop("context_text", None)
    return next_params


# ------------------------------------------------------------------
# 辅助: 设置 is_current
# ------------------------------------------------------------------

async def _set_current_asr(db: AsyncSession, patient_id: int, current_id: int):
    """将指定 ASR 结果设为当前, 同 patient 其他设为 False"""
    await db.execute(
        update(PatientAsrResult)
        .where(PatientAsrResult.patient_id == patient_id)
        .values(is_current=False)
    )
    await db.execute(
        update(PatientAsrResult)
        .where(PatientAsrResult.id == current_id)
        .values(is_current=True)
    )
    await db.commit()


async def _set_current_llm(db: AsyncSession, patient_id: int, current_id: int):
    await db.execute(
        update(PatientLlmResult)
        .where(PatientLlmResult.patient_id == patient_id)
        .values(is_current=False)
    )
    await db.execute(
        update(PatientLlmResult)
        .where(PatientLlmResult.id == current_id)
        .values(is_current=True)
    )
    await db.commit()


def _segments_text(segments) -> str:
    if not isinstance(segments, list):
        return ""
    return "\n".join(str(item.get("text") or "") for item in segments if isinstance(item, dict))


def _snapshot_params(config_snapshot) -> dict:
    if isinstance(config_snapshot, dict):
        return config_snapshot.get("params") or {}
    return {}


def _build_asr_integrity(record, audio_segment_count: int | None = None) -> dict:
    """ASR 完整性判断：同时支持分段模式和整段模式。

    分段模式可直接按已保存 seg_index 判断缺段；整段模式不把 1 段结果当缺段，
    只暴露文本长度/数字数量，供前端继续与历史最佳做横向比较。
    """
    if isinstance(record, dict):
        status = record.get("status")
        segments = record.get("segments") or []
        full_transcript = record.get("full_transcript") or ""
        params = _snapshot_params(record.get("config_snapshot"))
    else:
        status = getattr(record, "status", None)
        segments = getattr(record, "segments", None) or []
        full_transcript = getattr(record, "full_transcript", None) or ""
        params = _snapshot_params(getattr(record, "config_snapshot", None))

    text = full_transcript or _segments_text(segments)
    mode = str(params.get("audio_input_mode") or params.get("recognition_mode") or "segments")
    valid_segments = [
        item for item in segments
        if isinstance(item, dict) and str(item.get("text") or "").strip()
    ] if isinstance(segments, list) else []
    processed_segments = [
        item for item in segments
        if isinstance(item, dict) and str(item.get("seg_index", "")).isdigit()
    ] if isinstance(segments, list) else []
    empty_indices = sorted({
        int(item.get("seg_index"))
        for item in processed_segments
        if not str(item.get("text") or "").strip()
    })
    result_seg_count = len(valid_segments)
    processed_indices = sorted({
        int(item.get("seg_index"))
        for item in processed_segments
    })
    total = int(audio_segment_count or 0)
    is_segment_mode = mode in {"segments", "segment", "original_segments"} or not mode
    missing_indices: list[int] = []
    if total > 0 and is_segment_mode:
        missing_indices = [idx for idx in range(1, total + 1) if idx not in processed_indices]

    number_count = len(_NUMBER_RE.findall(text or ""))
    text_len = len(text or "")
    level = "complete"
    label = f"完整 {result_seg_count}/{total}" if total and is_segment_mode else "完整"
    score = 100
    reasons: list[str] = []
    effective_total = max(0, total - len(empty_indices)) if total and is_segment_mode else 0

    if status == "running":
        level = "running"
        label = "进行中"
        score = 0
    elif missing_indices:
        processed_count = len(processed_indices)
        level = "partial"
        label = f"部分转写 {processed_count}/{total}"
        score = max(30, round((processed_count / total) * 100))
        reasons.append(f"缺失分段：{','.join(map(str, missing_indices))}")
    elif empty_indices:
        level = "complete_with_empty"
        label = f"有效转写 {result_seg_count}/{effective_total}，空段 {len(empty_indices)}" if total and is_segment_mode else f"有空段 {len(empty_indices)}"
        score = 100
        reasons.append(f"空段/无有效语音：{','.join(map(str, empty_indices))}")
    elif status != "success":
        if text_len > 0 or result_seg_count > 0:
            level = "partial"
            label = f"部分转写 {result_seg_count}/{total}" if total and is_segment_mode else "部分转写"
            score = max(10, min(60, round((result_seg_count / total) * 100))) if total and is_segment_mode else 40
            reasons.append("任务失败但已保存部分分段")
        else:
            level = "failed"
            label = "失败 0/{}".format(total) if total and is_segment_mode else "失败"
            score = 0
            reasons.append("任务失败且无有效文本")
    elif text_len == 0:
        level = "failed"
        label = "无文本"
        score = 0
        reasons.append("ASR 未返回有效文本")

    return {
        "level": level,
        "label": label,
        "score": score,
        "audio_segment_count": total or None,
        "processed_segment_count": len(processed_indices),
        "result_segment_count": result_seg_count,
        "empty_segment_count": len(empty_indices),
        "empty_segment_indices": empty_indices,
        "missing_segment_indices": missing_indices,
        "text_length": text_len,
        "number_count": number_count,
        "audio_input_mode": mode,
        "reasons": reasons,
    }


async def cleanup_stale_asr_tasks(max_age_seconds: int = ASR_STALE_RUNNING_SECONDS) -> int:
    """Mark orphaned ASR running records as failed after app restarts.

    Background ASR tasks live only inside the current Python process. If the
    server is restarted while a task is running, the DB record can remain in
    `running` forever. Only records older than the threshold are touched, so a
    freshly started task in this process won't be changed.
    """
    cutoff = datetime.utcnow() - timedelta(seconds=max_age_seconds)
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            update(PatientAsrResult)
            .where(
                PatientAsrResult.status == "running",
                PatientAsrResult.created_at < cutoff,
            )
            .values(
                status="failed",
                error_message="ASR后台任务超时/服务重启前卡住，请重新发起识别",
                updated_at=datetime.utcnow(),
            )
        )
        await db.commit()
        return result.rowcount or 0


# ------------------------------------------------------------------
# ASR 接口
# ------------------------------------------------------------------

async def _prepare_asr_snapshot(
    db: AsyncSession,
    patient_id: int,
    asr_model_id: int,
    hotwords: Optional[str] = None,
    variant_name: Optional[str] = None,
    params_override: Optional[str] = None,
    source: Optional[str] = None,
    experiment_key: Optional[str] = None,
    config_hash: Optional[str] = None,
) -> dict:
    """读取 ASR 执行所需快照，避免后台任务持有 ORM 对象。"""
    patient_result = await db.execute(
        select(PatientRecord)
        .options(selectinload(PatientRecord.date_folder))
        .where(PatientRecord.id == patient_id)
    )
    patient = patient_result.scalar_one_or_none()
    if not patient:
        raise HTTPException(status_code=404, detail=f"患者 {patient_id} 不存在")

    segs_result = await db.execute(
        select(AudioSeg).where(AudioSeg.patient_id == patient_id).order_by(AudioSeg.seg_index)
    )
    raw_segs = [
        {"seg_index": s.seg_index, "file_path": s.file_path, "duration": s.duration}
        for s in segs_result.scalars().all()
    ]
    if not raw_segs:
        raise HTTPException(status_code=400, detail="该患者无录音文件")

    asr_model = await db.get(ModelConfig, asr_model_id)
    if not asr_model:
        raise HTTPException(status_code=404, detail="ASR 模型不存在")

    model_params = dict(asr_model.params or {})
    override_params = _parse_asr_params_override(params_override)
    if override_params:
        model_params.update(override_params)
    model_params = _apply_asr_feature_switches(model_params)

    parsed_hotwords = [w.strip() for w in hotwords.split(",")] if hotwords else None
    resolved_hotwords = resolve_hotwords(parsed_hotwords, model_params) if model_params.get("use_context_hotwords", True) is not False else None

    safe_variant_name = str(variant_name or "").strip()
    snap_asr_model_name = safe_variant_name[:100] if safe_variant_name else asr_model.name
    snap_source = (source or "normal").strip()[:50] or "normal"
    snap_experiment_key = (experiment_key or "").strip()[:100] or None
    snap_config_hash = (config_hash or "").strip()[:64] or None
    snap_config_snapshot = {
        "base_asr_model_id": asr_model_id,
        "provider": asr_model.provider,
        "model_name": asr_model.model_name,
        "params": model_params,
        "variant_name": snap_asr_model_name,
    }
    snap_asr_config = {
        "endpoint": asr_model.endpoint,
        "api_key": asr_model.api_key,
        "api_secret": asr_model.api_secret,
        "secret_key": asr_model.secret_key,
        "model_name": asr_model.model_name,
        "params": model_params,
    }
    snap_segs = build_asr_audio_inputs(raw_segs, model_params)
    logger.info(
        f"患者 {patient_id} ASR 准备: model={snap_asr_model_name}, "
        f"provider={asr_model.provider}, source={snap_source}, "
        f"audio_input_mode={model_params.get('audio_input_mode') or model_params.get('recognition_mode') or 'segments'}, "
        f"actual_inputs={len(snap_segs)}, "
        f"input_modes={sorted({str(seg.get('input_mode', 'segments')) for seg in snap_segs})}, "
        f"hotwords_count={len(resolved_hotwords or [])}, "
        f"config_hash={snap_config_hash or '-'}"
    )
    return {
        "patient_id": patient_id,
        "record_id": patient.record_id,
        "date": patient.date_folder.date if patient.date_folder else None,
        "asr_model_id": asr_model_id,
        "asr_model_name": snap_asr_model_name,
        "provider": asr_model.provider,
        "source": snap_source,
        "experiment_key": snap_experiment_key,
        "config_hash": snap_config_hash,
        "config_snapshot": snap_config_snapshot,
        "hotwords": resolved_hotwords or [],
        "asr_config": snap_asr_config,
        "segs": snap_segs,
    }


async def _run_asr_task(record_id_internal: int, snap: dict) -> None:
    """后台执行 ASR。客户端断开不影响任务继续运行。"""
    async with AsyncSessionLocal() as task_db:
        try:
            asr = create_asr(snap["provider"], **snap["asr_config"])
            start = time.time()
            asr_results = []
            timeout_seconds = int(
                (snap.get("asr_config") or {}).get("params", {}).get(
                    "task_timeout_seconds",
                    ASR_TRANSCRIBE_TIMEOUT_SECONDS,
                )
            )
            for seg in snap["segs"]:
                text = await asyncio.wait_for(
                    asr.transcribe(seg["file_path"], hotwords=snap["hotwords"]),
                    timeout=timeout_seconds,
                )
                text = text or ""
                asr_results.append({
                    "seg_index": seg["seg_index"],
                    "text": text,
                    "duration": seg["duration"],
                    "input_mode": seg.get("input_mode", "segments"),
                    "source_seg_count": seg.get("source_seg_count"),
                })

                record = await task_db.get(PatientAsrResult, record_id_internal)
                if record:
                    record.segments = list(asr_results)
                    record.duration_seconds = round(time.time() - start, 2)
                    await task_db.commit()

            full_transcript = "\n".join(r["text"] for r in asr_results)
            if not full_transcript.strip():
                raise ValueError("ASR 未返回有效转写文本")

            record = await task_db.get(PatientAsrResult, record_id_internal)
            if not record:
                return
            record.segments = list(asr_results)
            record.full_transcript = full_transcript
            record.duration_seconds = round(time.time() - start, 2)
            record.status = "success"
            record.error_message = None
            record.is_current = snap["source"] != "asr_optimization"
            await task_db.commit()

            if snap["source"] != "asr_optimization":
                await task_db.execute(
                    update(PatientAsrResult)
                    .where(
                        PatientAsrResult.patient_id == snap["patient_id"],
                        PatientAsrResult.id != record_id_internal,
                    )
                    .values(is_current=False)
                )
                await task_db.commit()
        except Exception as e:
            logger.error(f"患者 {snap.get('patient_id')} ASR 后台任务失败: {e}")
            try:
                record = await task_db.get(PatientAsrResult, record_id_internal)
                if record:
                    partial_text = _segments_text(record.segments)
                    if partial_text.strip():
                        record.full_transcript = partial_text
                        record.status = "partial"
                    else:
                        record.status = "failed"
                    record.error_message = str(e)
                    record.duration_seconds = round(time.time() - record.created_at.timestamp(), 2) if record.created_at else None
                    await task_db.commit()
            except Exception:
                pass


def _schedule_asr_task(record_id_internal: int, snap: dict) -> None:
    """将 ASR 任务挂到当前事件循环，避免请求断开导致任务生命周期不清晰。"""
    task = asyncio.create_task(_run_asr_task(record_id_internal, snap))
    _RUNNING_ASR_TASKS.add(task)

    def _cleanup(done_task: asyncio.Task) -> None:
        _RUNNING_ASR_TASKS.discard(done_task)
        try:
            done_task.result()
        except Exception as exc:
            logger.error(f"ASR 后台任务未捕获异常: {exc}")

    task.add_done_callback(_cleanup)


@router.post("/{patient_id}/asr/tasks")
async def start_patient_asr_task(
    patient_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
):
    """创建后台 ASR 任务并立即返回 running 记录。"""
    asr_model_id = body.get("asr_model_id")
    if not asr_model_id:
        raise HTTPException(status_code=400, detail="缺少 asr_model_id")
    snap = await _prepare_asr_snapshot(
        db=db,
        patient_id=patient_id,
        asr_model_id=int(asr_model_id),
        hotwords=body.get("hotwords"),
        variant_name=body.get("variant_name"),
        params_override=json.dumps(body.get("params_override"), ensure_ascii=False) if isinstance(body.get("params_override"), dict) else body.get("params_override"),
        source=body.get("source"),
        experiment_key=body.get("experiment_key"),
        config_hash=body.get("config_hash"),
    )
    record = PatientAsrResult(
        patient_id=snap["patient_id"],
        record_id=snap["record_id"],
        date=snap["date"],
        asr_model_id=snap["asr_model_id"],
        asr_model_name=snap["asr_model_name"],
        provider=snap["provider"],
        source=snap["source"],
        experiment_key=snap["experiment_key"],
        config_hash=snap["config_hash"],
        config_snapshot=snap["config_snapshot"],
        hotwords=snap["hotwords"],
        status="running",
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)
    _schedule_asr_task(record.id, snap)
    return {"stage": "started", "result_id": record.id, **_asr_response(record)}


@router.get("/{patient_id}/asr/tasks/{result_id}")
async def get_patient_asr_task(
    patient_id: int,
    result_id: int,
    db: AsyncSession = Depends(get_db),
):
    record = await db.get(PatientAsrResult, result_id)
    if not record or record.patient_id != patient_id:
        raise HTTPException(status_code=404, detail="ASR 任务不存在")
    audio_count = await _get_audio_segment_count(db, patient_id)
    return _asr_response(record, audio_count)


async def _get_audio_segment_count(db: AsyncSession, patient_id: int) -> int:
    result = await db.execute(select(AudioSeg).where(AudioSeg.patient_id == patient_id))
    return len(result.scalars().all())


async def _build_repair_snapshot(db: AsyncSession, record: PatientAsrResult) -> tuple[dict, list[dict]]:
    snapshot = record.config_snapshot or {}
    if not isinstance(snapshot, dict):
        snapshot = {}
    params = dict((snapshot.get("params") or {}))
    mode = str(params.get("audio_input_mode") or params.get("recognition_mode") or "segments")
    if mode not in {"segments", "segment", "original_segments"}:
        raise HTTPException(status_code=400, detail="当前 ASR 结果不是原始分段模式，不能按缺失段补跑，请使用完整重跑")

    asr_model_id = int(snapshot.get("base_asr_model_id") or record.asr_model_id or 0)
    asr_model = await db.get(ModelConfig, asr_model_id)
    if not asr_model:
        raise HTTPException(status_code=404, detail="原 ASR 模型不存在，无法补跑")

    segs_result = await db.execute(
        select(AudioSeg).where(AudioSeg.patient_id == record.patient_id).order_by(AudioSeg.seg_index)
    )
    raw_segs = [
        {"seg_index": s.seg_index, "file_path": s.file_path, "duration": s.duration}
        for s in segs_result.scalars().all()
    ]
    existing = {
        int(item.get("seg_index"))
        for item in (record.segments or [])
        if isinstance(item, dict) and str(item.get("seg_index", "")).isdigit()
    }
    missing_raw = [seg for seg in raw_segs if int(seg["seg_index"]) not in existing]
    if not missing_raw:
        raise HTTPException(status_code=400, detail="没有需要补跑的缺失分段")

    params["audio_input_mode"] = "segments"
    # 补跑时强制关闭分句，避开豆包 confidence=NaN / utterances 解析问题。
    if record.provider == "volcengine":
        params["show_utterances"] = False
    params = _apply_asr_feature_switches(params)
    hotwords = record.hotwords or []
    if params.get("use_context_hotwords", True) is False:
        hotwords = []
    snap = {
        "provider": record.provider,
        "hotwords": hotwords,
        "asr_config": {
            "endpoint": asr_model.endpoint,
            "api_key": asr_model.api_key,
            "api_secret": asr_model.api_secret,
            "secret_key": asr_model.secret_key,
            "model_name": asr_model.model_name,
            "params": params,
        },
    }
    return snap, missing_raw


@router.post("/{patient_id}/asr-results/{result_id}/repair-missing-segments")
async def repair_patient_asr_missing_segments(
    patient_id: int,
    result_id: int,
    db: AsyncSession = Depends(get_db),
):
    """补跑分段模式 ASR 的缺失段，并合并回原 ASR 结果。"""
    record = await db.get(PatientAsrResult, result_id)
    if not record or record.patient_id != patient_id:
        raise HTTPException(status_code=404, detail="ASR 记录不存在")

    snap, missing_raw = await _build_repair_snapshot(db, record)
    asr = create_asr(snap["provider"], **snap["asr_config"])
    timeout_seconds = int(
        (snap.get("asr_config") or {}).get("params", {}).get(
            "task_timeout_seconds",
            ASR_TRANSCRIBE_TIMEOUT_SECONDS,
        )
    )
    existing_segments = [
        item for item in (record.segments or [])
        if isinstance(item, dict) and str(item.get("seg_index", "")).isdigit()
    ]
    repaired_segments: list[dict] = []
    failed_segments: list[dict] = []

    for seg in missing_raw:
        try:
            text = await asyncio.wait_for(
                asr.transcribe(seg["file_path"], hotwords=snap["hotwords"]),
                timeout=timeout_seconds,
            )
            repaired_segments.append({
                "seg_index": seg["seg_index"],
                "text": text or "",
                "duration": seg.get("duration"),
                "input_mode": "segments",
                "repaired": True,
                "empty": not bool((text or "").strip()),
                "empty_reason": "asr_return_empty" if not (text or "").strip() else None,
            })
        except Exception as exc:
            failed_segments.append({"seg_index": seg["seg_index"], "error": str(exc)})

    merged = existing_segments + repaired_segments
    merged.sort(key=lambda item: int(item.get("seg_index") or 0))
    record.segments = merged
    record.full_transcript = "\n".join(str(item.get("text") or "") for item in merged)
    audio_count = await _get_audio_segment_count(db, patient_id)
    existing_indices = {
        int(item.get("seg_index"))
        for item in merged
        if isinstance(item, dict) and str(item.get("seg_index", "")).isdigit()
    }
    missing_after = [idx for idx in range(1, audio_count + 1) if idx not in existing_indices]
    if missing_after:
        record.status = "partial"
        record.error_message = f"补跑后仍缺失分段: {','.join(map(str, missing_after))}"
    else:
        record.status = "success"
        record.error_message = None
    if failed_segments:
        record.error_message = f"部分分段补跑失败: {json.dumps(failed_segments, ensure_ascii=False)}"
    await db.commit()
    await db.refresh(record)
    return {
        **_asr_response(record, audio_count),
        "repaired_segments": [item["seg_index"] for item in repaired_segments],
        "failed_segments": failed_segments,
    }

@router.get("/{patient_id}/asr/stream")
async def patient_asr_stream(
    patient_id: int,
    asr_model_id: int,
    hotwords: Optional[str] = None,
    variant_name: Optional[str] = Query(None),
    params_override: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    experiment_key: Optional[str] = Query(None),
    config_hash: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """患者级 SSE 流式 ASR, 结果持久化到 patient_asr_results

    前置阶段只读取必要快照 (plain dict), 避免跨 stream 长期持有 ORM 对象。
    event_generator 内部使用独立 AsyncSessionLocal, 确保流式期间 DB 操作可靠。
    """
    # === 前置阶段: 读取必要快照 (不把 ORM 对象传进 generator) ===
    patient_result = await db.execute(
        select(PatientRecord)
        .options(selectinload(PatientRecord.date_folder))
        .where(PatientRecord.id == patient_id)
    )
    patient = patient_result.scalar_one_or_none()
    if not patient:
        raise HTTPException(status_code=404, detail=f"患者 {patient_id} 不存在")

    segs_result = await db.execute(
        select(AudioSeg).where(AudioSeg.patient_id == patient_id).order_by(AudioSeg.seg_index)
    )
    raw_segs = [
        {"seg_index": s.seg_index, "file_path": s.file_path, "duration": s.duration}
        for s in segs_result.scalars().all()
    ]
    if not raw_segs:
        raise HTTPException(status_code=400, detail="该患者无录音文件")

    asr_model = await db.get(ModelConfig, asr_model_id)
    if not asr_model:
        raise HTTPException(status_code=404, detail="ASR 模型不存在")

    model_params = dict(asr_model.params or {})
    override_params = _parse_asr_params_override(params_override)
    if override_params:
        model_params.update(override_params)
    model_params = _apply_asr_feature_switches(model_params)

    parsed_hotwords = [w.strip() for w in hotwords.split(",")] if hotwords else None
    resolved_hotwords = resolve_hotwords(parsed_hotwords, model_params) if model_params.get("use_context_hotwords", True) is not False else None

    # 快照所有需要的纯数据 (不传 ORM 对象进 generator)
    snap_patient_id = patient_id
    snap_record_id = patient.record_id
    snap_date = patient.date_folder.date if patient.date_folder else None
    snap_asr_model_id = asr_model_id
    safe_variant_name = str(variant_name or "").strip()
    snap_asr_model_name = safe_variant_name[:100] if safe_variant_name else asr_model.name
    snap_provider = asr_model.provider
    snap_source = (source or "normal").strip()[:50] or "normal"
    snap_experiment_key = (experiment_key or "").strip()[:100] or None
    snap_config_hash = (config_hash or "").strip()[:64] or None
    snap_config_snapshot = {
        "base_asr_model_id": asr_model_id,
        "provider": asr_model.provider,
        "model_name": asr_model.model_name,
        "params": model_params,
        "variant_name": snap_asr_model_name,
    }
    snap_hotwords = resolved_hotwords or []
    snap_asr_config = {
        "endpoint": asr_model.endpoint,
        "api_key": asr_model.api_key,
        "api_secret": asr_model.api_secret,
        "secret_key": asr_model.secret_key,
        "model_name": asr_model.model_name,
        "params": model_params,
    }
    snap_segs = build_asr_audio_inputs(raw_segs, model_params)
    logger.info(
        f"患者 {snap_patient_id} ASR 准备: model={snap_asr_model_name}, "
        f"provider={snap_provider}, source={snap_source}, "
        f"audio_input_mode={model_params.get('audio_input_mode') or model_params.get('recognition_mode') or 'segments'}, "
        f"actual_inputs={len(snap_segs)}, "
        f"input_modes={sorted({str(seg.get('input_mode', 'segments')) for seg in snap_segs})}, "
        f"hotwords_count={len(snap_hotwords)}, "
        f"config_hash={snap_config_hash or '-'}"
    )

    async def event_generator():
        # === 独立 session, 不依赖请求级 db ===
        async with AsyncSessionLocal() as stream_db:
            record_id_internal = None
            try:
                # 创建 running 记录
                record = PatientAsrResult(
                    patient_id=snap_patient_id,
                    record_id=snap_record_id,
                    date=snap_date,
                    asr_model_id=snap_asr_model_id,
                    asr_model_name=snap_asr_model_name,
                    provider=snap_provider,
                    source=snap_source,
                    experiment_key=snap_experiment_key,
                    config_hash=snap_config_hash,
                    config_snapshot=snap_config_snapshot,
                    hotwords=snap_hotwords,
                    status="running",
                )
                stream_db.add(record)
                await stream_db.commit()
                await stream_db.refresh(record)
                record_id_internal = record.id

                yield f"event: progress\ndata: {json.dumps({'stage': 'progress', 'total': len(snap_segs), 'started': True}, ensure_ascii=False)}\n\n"

                # 使用独立 asr 实例
                asr = create_asr(snap_provider, **snap_asr_config)
                start = time.time()
                asr_results = []

                for seg in snap_segs:
                    yield f"event: progress\ndata: {json.dumps({'stage': 'segment_start', 'seg_index': seg['seg_index'], 'total': len(snap_segs), 'input_mode': seg.get('input_mode')}, ensure_ascii=False)}\n\n"

                    text = await asr.transcribe(seg["file_path"], hotwords=snap_hotwords)
                    text = text or ""
                    asr_results.append({
                        "seg_index": seg["seg_index"],
                        "text": text,
                        "duration": seg["duration"],
                        "input_mode": seg.get("input_mode", "segments"),
                        "source_seg_count": seg.get("source_seg_count"),
                    })

                    yield f"event: segment\ndata: {json.dumps({'stage': 'segment', 'seg_index': seg['seg_index'], 'text': text, 'duration': seg['duration']}, ensure_ascii=False)}\n\n"

                full_transcript = "\n".join(r["text"] for r in asr_results)
                if not full_transcript.strip():
                    raise ValueError("ASR 未返回有效转写文本")
                duration_val = round(time.time() - start, 2)

                # 重新获取记录并更新 (避免悬挂 ORM)
                record = await stream_db.get(PatientAsrResult, record_id_internal)
                record.segments = list(asr_results)
                record.full_transcript = full_transcript
                record.duration_seconds = duration_val
                record.status = "success"
                record.is_current = snap_source != "asr_optimization"
                await stream_db.commit()

                # 普通业务结果才切换 current；优化评估结果不污染数据管理当前结果
                if snap_source != "asr_optimization":
                    await stream_db.execute(
                        update(PatientAsrResult)
                        .where(
                            PatientAsrResult.patient_id == snap_patient_id,
                            PatientAsrResult.id != record_id_internal,
                        )
                        .values(is_current=False)
                    )
                    await stream_db.commit()

                # commit 成功后再发送 complete
                yield f"event: complete\ndata: {json.dumps({'stage': 'complete', 'result_id': record_id_internal, **_asr_response(record)}, ensure_ascii=False)}\n\n"

            except asyncio.CancelledError:
                # 客户端中断 (关闭页面 / 网络断开)
                logger.warning(f"患者 {snap_patient_id} ASR SSE 连接中断")
                try:
                    if record_id_internal:
                        record = await stream_db.get(PatientAsrResult, record_id_internal)
                        record.status = "failed"
                        record.error_message = "SSE 连接中断或客户端取消"
                        await stream_db.commit()
                except Exception:
                    pass
                yield f"event: error\ndata: {json.dumps({'stage': 'error', 'message': '连接中断'}, ensure_ascii=False)}\n\n"

            except Exception as e:
                logger.error(f"患者 {snap_patient_id} ASR 失败: {e}")
                try:
                    if record_id_internal:
                        record = await stream_db.get(PatientAsrResult, record_id_internal)
                        partial_text = _segments_text(record.segments)
                        if partial_text.strip():
                            record.full_transcript = partial_text
                            record.status = "partial"
                        else:
                            record.status = "failed"
                        record.error_message = str(e)
                        await stream_db.commit()
                except Exception:
                    pass
                yield f"event: error\ndata: {json.dumps({'stage': 'error', 'message': str(e)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/{patient_id}/asr-results")
async def list_patient_asr_results(patient_id: int, db: AsyncSession = Depends(get_db)):
    """返回某患者所有 ASR 历史"""
    result = await db.execute(
        select(PatientAsrResult)
        .where(
            PatientAsrResult.patient_id == patient_id,
            or_(PatientAsrResult.source.is_(None), PatientAsrResult.source != "asr_optimization"),
        )
        .order_by(PatientAsrResult.created_at.desc())
    )
    audio_count = await _get_audio_segment_count(db, patient_id)
    return [_asr_response(r, audio_count) for r in result.scalars().all()]


@router.get("/asr-results/batch")
async def list_patient_asr_results_batch(
    patient_ids: str = Query("", description="逗号分隔的检查记录 ID"),
    db: AsyncSession = Depends(get_db),
):
    """批量返回检查记录 ASR 历史，用于 ASR 模型横向对比页面。"""
    ids: list[int] = []
    for item in patient_ids.split(","):
        item = item.strip()
        if not item:
            continue
        try:
            ids.append(int(item))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"无效检查记录 ID: {item}")

    if not ids:
        return {}

    seg_result = await db.execute(select(AudioSeg.patient_id, AudioSeg.id).where(AudioSeg.patient_id.in_(ids)))
    audio_counts: dict[int, int] = {}
    for patient_id, _seg_id in seg_result.all():
        audio_counts[int(patient_id)] = audio_counts.get(int(patient_id), 0) + 1

    result = await db.execute(
        select(PatientAsrResult)
        .where(PatientAsrResult.patient_id.in_(ids))
        .order_by(
            PatientAsrResult.patient_id.asc(),
            PatientAsrResult.asr_model_id.asc(),
            PatientAsrResult.created_at.desc(),
            PatientAsrResult.id.desc(),
        )
    )

    output: dict[str, list[dict]] = {str(i): [] for i in ids}
    for row in result.scalars().all():
        output.setdefault(str(row.patient_id), []).append(_asr_response(row, audio_counts.get(int(row.patient_id), 0)))
    return output


@router.get("/{patient_id}/asr-current")
async def get_patient_asr_current(patient_id: int, db: AsyncSession = Depends(get_db)):
    """返回默认展示的 ASR 结果：以最新一次为准，兼容历史 is_current 字段。"""
    result = await db.execute(
        select(PatientAsrResult)
        .where(
            PatientAsrResult.patient_id == patient_id,
            or_(PatientAsrResult.source.is_(None), PatientAsrResult.source != "asr_optimization"),
        )
        .order_by(PatientAsrResult.created_at.desc(), PatientAsrResult.id.desc())
    )
    record = result.scalars().first()
    audio_count = await _get_audio_segment_count(db, patient_id) if record else 0
    return _asr_response(record, audio_count) if record else None


def _asr_response(r: PatientAsrResult, audio_segment_count: int | None = None) -> dict:
    """构建兼容前端旧字段的响应"""
    return {
        "id": r.id,
        "exam_record_id": r.patient_id,  # patient_id 实际是 exam_record_id
        "patient_id": r.patient_id,      # 保留兼容
        "record_id": r.record_id,
        "date": r.date,
        "asr_model_id": r.asr_model_id,
        "model_name": r.asr_model_name or "",  # 前端旧字段
        "full_model_name": r.asr_model_name,
        "provider": r.provider,
        "source": r.source or "normal",
        "experiment_key": r.experiment_key,
        "config_hash": r.config_hash,
        "config_snapshot": r.config_snapshot,
        "segments": r.segments,
        "full_transcript": r.full_transcript,
        "duration_seconds": r.duration_seconds,
        "status": r.status,
        "error_message": r.error_message,
        "asr_integrity": _build_asr_integrity(r, audio_segment_count),
        "is_current": r.is_current,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


def _asr_reference_response(r: AsrReferenceTranscript) -> dict:
    """标准 ASR 文本响应。"""
    return {
        "id": r.id,
        "patient_id": r.patient_id,
        "record_id": r.record_id,
        "date": r.date,
        "base_asr_result_id": r.base_asr_result_id,
        "base_asr_model_name": r.base_asr_model_name,
        "base_config_hash": r.base_config_hash,
        "reference_text": r.reference_text or "",
        "reference_annotations": r.reference_annotations or [],
        "note": r.note or "",
        "is_current": r.is_current,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
    }


def _normalize_reference_annotations(value, text_length: int) -> list[dict]:
    """清洗标准 ASR 文本标注，避免前端传入越界或无效数据。"""
    if not isinstance(value, list):
        return []
    allowed_types = {"red", "orange", "green"}
    cleaned: list[dict] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        try:
            start = int(item.get("start"))
            end = int(item.get("end"))
        except (TypeError, ValueError):
            continue
        mark_type = str(item.get("type") or "red")
        if mark_type not in allowed_types:
            mark_type = "red"
        start = max(0, min(start, text_length))
        end = max(0, min(end, text_length))
        if end <= start:
            continue
        note = str(item.get("note") or "").strip()
        cleaned.append({
            "start": start,
            "end": end,
            "type": mark_type,
            "note": note,
        })
    return sorted(cleaned, key=lambda row: (row["start"], row["end"]))


@router.get("/{patient_id}/asr-reference")
async def get_patient_asr_reference(patient_id: int, db: AsyncSession = Depends(get_db)):
    """返回检查记录当前标准 ASR 文本。"""
    result = await db.execute(
        select(AsrReferenceTranscript)
        .where(AsrReferenceTranscript.patient_id == patient_id, AsrReferenceTranscript.is_current == True)
        .order_by(AsrReferenceTranscript.updated_at.desc(), AsrReferenceTranscript.id.desc())
    )
    row = result.scalars().first()
    return _asr_reference_response(row) if row else None


@router.get("/asr-references/batch")
async def list_patient_asr_references_batch(
    patient_ids: str = Query("", description="逗号分隔的检查记录 ID"),
    db: AsyncSession = Depends(get_db),
):
    """批量返回检查记录标准 ASR 文本，用于优化评估左侧概览。"""
    ids: list[int] = []
    for item in patient_ids.split(","):
        item = item.strip()
        if not item:
            continue
        try:
            ids.append(int(item))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"无效检查记录 ID: {item}")

    if not ids:
        return {}

    result = await db.execute(
        select(AsrReferenceTranscript)
        .where(AsrReferenceTranscript.patient_id.in_(ids), AsrReferenceTranscript.is_current == True)
        .order_by(AsrReferenceTranscript.patient_id.asc(), AsrReferenceTranscript.updated_at.desc(), AsrReferenceTranscript.id.desc())
    )
    output: dict[str, dict] = {}
    for row in result.scalars().all():
        key = str(row.patient_id)
        if key not in output:
            output[key] = _asr_reference_response(row)
    return output


@router.put("/{patient_id}/asr-reference")
async def upsert_patient_asr_reference(
    patient_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
):
    """保存/更新检查记录标准 ASR 文本。"""
    reference_text = str(body.get("reference_text") or "").strip()
    if not reference_text:
        raise HTTPException(status_code=400, detail="标准 ASR 文本不能为空")

    patient_result = await db.execute(
        select(PatientRecord)
        .options(selectinload(PatientRecord.date_folder))
        .where(PatientRecord.id == patient_id)
    )
    patient = patient_result.scalars().first()
    if not patient:
        raise HTTPException(status_code=404, detail="检查记录不存在")

    base_asr_result = None
    base_asr_result_id = body.get("base_asr_result_id")
    if base_asr_result_id:
        try:
            base_asr_result_id = int(base_asr_result_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="base_asr_result_id 无效")
        base_asr_result = await db.get(PatientAsrResult, base_asr_result_id)
        if not base_asr_result or base_asr_result.patient_id != patient_id:
            raise HTTPException(status_code=404, detail="底稿 ASR 结果不存在或不属于该检查记录")

    date = patient.date_folder.date if patient.date_folder else None

    result = await db.execute(
        select(AsrReferenceTranscript)
        .where(AsrReferenceTranscript.patient_id == patient_id, AsrReferenceTranscript.is_current == True)
        .order_by(AsrReferenceTranscript.updated_at.desc(), AsrReferenceTranscript.id.desc())
    )
    row = result.scalars().first()
    now = datetime.utcnow()
    note = str(body.get("note") or "").strip() or None
    reference_annotations = _normalize_reference_annotations(body.get("reference_annotations"), len(reference_text))

    if row:
        row.record_id = patient.record_id
        row.date = date
        row.reference_text = reference_text
        row.reference_annotations = reference_annotations
        row.note = note
        row.base_asr_result_id = base_asr_result.id if base_asr_result else row.base_asr_result_id
        row.base_asr_model_name = (base_asr_result.asr_model_name if base_asr_result else row.base_asr_model_name)
        row.base_config_hash = (base_asr_result.config_hash if base_asr_result else row.base_config_hash)
        row.updated_at = now
    else:
        row = AsrReferenceTranscript(
            patient_id=patient_id,
            record_id=patient.record_id,
            date=date,
            base_asr_result_id=base_asr_result.id if base_asr_result else None,
            base_asr_model_name=base_asr_result.asr_model_name if base_asr_result else None,
            base_config_hash=base_asr_result.config_hash if base_asr_result else None,
            reference_text=reference_text,
            reference_annotations=reference_annotations,
            note=note,
            is_current=True,
            created_at=now,
            updated_at=now,
        )
        db.add(row)

    await db.commit()
    await db.refresh(row)
    return _asr_reference_response(row)


def _llm_response(r: PatientLlmResult) -> dict:
    from app.services.parser import normalize_structured_result
    structured_result = normalize_structured_result(r.structured_result)
    evaluation = r.evaluation
    accuracy_without_remark = r.accuracy
    if isinstance(evaluation, dict):
        accuracy_without_remark = evaluation.get("accuracy", accuracy_without_remark)
    asr_model_name = ""
    if r.asr_result is not None:
        # 已 preload
        asr_model_name = r.asr_result.asr_model_name or ""
    elif r.asr_result_id:
        # 懒加载回退 (在 async 上下文中需要 spawn)
        pass
    return {
        "id": r.id,
        "exam_record_id": r.patient_id,  # patient_id 实际是 exam_record_id
        "patient_id": r.patient_id,      # 保留兼容
        "asr_result_id": r.asr_result_id,
        "asr_model_name": asr_model_name,
        "llm_model_id": r.llm_model_id,
        "model_name": r.llm_model_name or "",  # 前端旧字段
        "full_model_name": r.llm_model_name,
        "prompt_template_id": r.prompt_template_id,
        "prompt_template_name": r.prompt_template_name,
        "prompt_version": r.prompt_version,
        "prompt_content": r.prompt_content,
        "prompt_len": len(r.prompt_content) if r.prompt_content else 0,
        "source": r.source or "normal",
        "experiment_key": r.experiment_key,
        "structured": structured_result,        # 前端旧字段
        "structured_result": structured_result,
        "summary": r.summary_text,                # 前端旧字段
        "summary_text": r.summary_text,
        "raw_text": r.raw_output,                 # 前端旧字段
        "raw_output": r.raw_output,
        "evaluation": evaluation,
        "accuracy": accuracy_without_remark,
        "accuracy_without_remark": accuracy_without_remark,
        "accuracy_with_remark": None,
        "status": r.status,
        "error_message": r.error_message,
        "is_current": r.is_current,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


@router.put("/{patient_id}/asr-results/{result_id}/current")
async def set_patient_asr_current(
    patient_id: int,
    result_id: int,
    db: AsyncSession = Depends(get_db),
):
    """切换当前 ASR 结果"""
    record = await db.get(PatientAsrResult, result_id)
    if not record or record.patient_id != patient_id:
        raise HTTPException(status_code=404, detail="记录不存在")
    await _set_current_asr(db, patient_id, result_id)
    return {"ok": True}


# ------------------------------------------------------------------
# LLM 接口
# ------------------------------------------------------------------

@router.post("/{patient_id}/llm/run")
async def patient_llm_run(
    patient_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
):
    """患者级 LLM 结构化提取, 保存到 patient_llm_results"""
    llm_model_id = body.get("llm_model_id")
    asr_result_id = body.get("asr_result_id")  # 可选, 默认当前
    prompt_content = body.get("prompt_content")
    prompt_template_id = body.get("prompt_template_id")  # 可选
    source = (body.get("source") or "normal").strip()[:50] or "normal"
    experiment_key = (body.get("experiment_key") or "").strip()[:100] or None

    # 如果没有传入 prompt_content,尝试从模板加载
    if not prompt_content and prompt_template_id:
        from app.models import PromptTemplate
        tmpl = await db.get(PromptTemplate, prompt_template_id)
        if tmpl:
            prompt_content = tmpl.content
            # 也保存模板ID用于后续记录
            if not prompt_template_id:
                prompt_template_id = tmpl.id

    patient = await db.get(PatientRecord, patient_id)
    if not patient:
        raise HTTPException(status_code=404, detail=f"患者 {patient_id} 不存在")

    # 确定 ASR 结果
    asr_record = None
    if asr_result_id:
        asr_record = await db.get(PatientAsrResult, asr_result_id)
    else:
        # 默认取当前
        result = await db.execute(
            select(PatientAsrResult)
            .where(PatientAsrResult.patient_id == patient_id, PatientAsrResult.is_current == True)
        )
        asr_record = result.scalar_one_or_none()

    transcript = asr_record.full_transcript if asr_record else ""
    if not transcript.strip():
        raise HTTPException(status_code=400, detail="无 ASR 转写文本可用")

    # 读取 LLM 模型并校验配置完整性
    if not llm_model_id:
        raise HTTPException(status_code=400, detail="请提供 llm_model_id")
    llm_model = await db.get(ModelConfig, llm_model_id)
    if not llm_model:
        raise HTTPException(status_code=404, detail="LLM 模型不存在")

    # 校验 LLM 配置是否完整（避免后续 401 / 500 浪费资源）
    from app.models import ModelConfig as MC
    if llm_model.model_type != "llm":
        raise HTTPException(status_code=400, detail=f"{llm_model.name} 不是 LLM 模型")
    if llm_model.status != "active":
        raise HTTPException(status_code=400, detail=f"{llm_model.name} 当前未激活（status={llm_model.status}），请先在模型配置中启用")
    missing = []
    if not llm_model.endpoint:
        missing.append("endpoint")
    if not llm_model.api_key:
        missing.append("API Key")
    if not llm_model.model_name:
        missing.append("model_name")
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"{llm_model.name} 未配置: {', '.join(missing)}。请先在「模型配置」中补全信息后再执行。",
        )

    # 尝试获取提示词模板信息
    prompt_template_name = None
    if prompt_template_id:
        from app.models import PromptTemplate
        tmpl = await db.get(PromptTemplate, prompt_template_id)
        if tmpl:
            prompt_template_name = tmpl.name

    # 创建 running 记录
    record = PatientLlmResult(
        patient_id=patient_id,
        asr_result_id=asr_record.id if asr_record else None,
        llm_model_id=llm_model_id,
        llm_model_name=llm_model.name,
        prompt_template_id=prompt_template_id,
        prompt_template_name=prompt_template_name,
        prompt_content=prompt_content,
        source=source,
        experiment_key=experiment_key,
        status="running",
    )
    # 显式设置 asr_result 关系,避免后续 _llm_response 访问时触发懒加载
    if asr_record:
        record.asr_result = asr_record
    db.add(record)
    await db.commit()
    await db.refresh(record)

    executor = TestExecutor()
    try:
        llm_result = await executor.execute_llm(
            transcript=transcript,
            llm_provider=llm_model.provider,
            llm_config={
                "endpoint": llm_model.endpoint,
                "api_key": llm_model.api_key,
                "api_secret": llm_model.api_secret,
                "model_name": llm_model.model_name,
                "params": llm_model.params or {},
            },
            prompt_template=prompt_content,
        )
        from app.services.parser import normalize_structured_result
        record.structured_result = normalize_structured_result(llm_result["structured_result"])
        record.raw_output = llm_result["llm_raw_output"]
        record.prompt_content = prompt_content  # 保存实际使用的提示词
        record.status = "success"

        # summary_text 优先取 structured 中的 summary 字段
        structured = record.structured_result or {}
        if structured.get("summary"):
            record.summary_text = str(structured["summary"])
        elif llm_result.get("summary_text"):
            record.summary_text = llm_result["summary_text"]
        else:
            record.summary_text = ""

        # 评估
        if record.structured_result:
            from sqlalchemy import select as sa_select
            from app.models import BUltraResult
            from app.services.parser import evaluate_result
            gt_r = await db.execute(
                sa_select(BUltraResult).where(BUltraResult.patient_id == patient_id)
            )
            gt = gt_r.scalar_one_or_none()
            if gt:
                evaluation = evaluate_result(
                    identified=record.structured_result,
                    ground_truth={
                        "right_follicle_total": gt.right_follicle_total,
                        "left_follicle_total": gt.left_follicle_total,
                        "right_follicles": gt.right_follicles,
                        "left_follicles": gt.left_follicles,
                        "endometrium_thickness": gt.endometrium_thickness,
                        "endometrium_type": gt.endometrium_type,
                        "right_ovary_length": gt.right_ovary_length,
                        "right_ovary_width": gt.right_ovary_width,
                        "left_ovary_length": gt.left_ovary_length,
                        "left_ovary_width": gt.left_ovary_width,
                    },
                )
                record.evaluation = evaluation
                record.accuracy = evaluation.get("accuracy")

        await db.commit()
        if source != "asr_optimization":
            await _set_current_llm(db, patient_id, record.id)
    except Exception as e:
        logger.error(f"患者 {patient_id} LLM 失败: {e}")
        record.status = "failed"
        record.error_message = str(e)
        await db.commit()
        raise HTTPException(status_code=500, detail=str(e))

    return _llm_response(record)


@router.get("/{patient_id}/llm-results")
async def list_patient_llm_results(
    patient_id: int,
    include_optimization: bool = Query(False, description="是否包含优化评估来源的 LLM 结果"),
    db: AsyncSession = Depends(get_db),
):
    query = (
        select(PatientLlmResult)
        .options(selectinload(PatientLlmResult.asr_result))
        .where(PatientLlmResult.patient_id == patient_id)
    )
    if not include_optimization:
        query = query.where(or_(PatientLlmResult.source.is_(None), PatientLlmResult.source != "asr_optimization"))
    result = await db.execute(query.order_by(PatientLlmResult.created_at.desc()))
    return [_llm_response(r) for r in result.scalars().all()]


@router.get("/{patient_id}/llm-current")
async def get_patient_llm_current(patient_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(PatientLlmResult)
        .options(selectinload(PatientLlmResult.asr_result))
        .where(PatientLlmResult.patient_id == patient_id)
        .order_by(PatientLlmResult.created_at.desc(), PatientLlmResult.id.desc())
    )
    record = result.scalars().first()
    return _llm_response(record) if record else None


@router.put("/{patient_id}/llm-results/{result_id}/current")
async def set_patient_llm_current(
    patient_id: int,
    result_id: int,
    db: AsyncSession = Depends(get_db),
):
    record = await db.get(PatientLlmResult, result_id)
    if not record or record.patient_id != patient_id:
        raise HTTPException(status_code=404, detail="记录不存在")
    await _set_current_llm(db, patient_id, result_id)
    return {"ok": True}


@router.delete("/{patient_id}/llm-results")
async def clear_patient_llm_results(
    patient_id: int,
    db: AsyncSession = Depends(get_db),
):
    """清空当前检查记录的全部 LLM 历史记录"""
    from sqlalchemy import delete

    patient = await db.get(PatientRecord, patient_id)
    if not patient:
        raise HTTPException(status_code=404, detail=f"检查记录 {patient_id} 不存在")

    result = await db.execute(
        delete(PatientLlmResult).where(PatientLlmResult.patient_id == patient_id)
    )
    deleted = result.rowcount
    await db.commit()
    return {"ok": True, "deleted": deleted}


@router.get("/{patient_id}/llm-results/export")
async def export_patient_llm_results(
    patient_id: int,
    db: AsyncSession = Depends(get_db),
):
    """导出当前检查记录的全部 LLM 历史 + ASR + 真实B超 + 提示词模板 为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from sqlalchemy import select
    from app.models import DateFolder, BUltraResult, PromptTemplate, PatientAsrResult
    from app.services.parser import normalize_structured_result

    result = await db.execute(
        select(PatientLlmResult, PatientRecord, DateFolder)
        .join(PatientRecord, PatientLlmResult.patient_id == PatientRecord.id)
        .outerjoin(DateFolder, PatientRecord.date_folder_id == DateFolder.id)
        .where(PatientLlmResult.patient_id == patient_id)
        .order_by(PatientLlmResult.created_at.asc())
    )
    rows = result.all()

    if not rows:
        raise HTTPException(status_code=404, detail="无 LLM 结果")

    patient = rows[0][1] if rows else None
    date_folder = rows[0][2] if rows else None
    record_id = patient.record_id if patient else f"exam_{patient_id}"
    date_str = date_folder.date if date_folder else ""

    # 获取关联数据
    gt_result = await db.execute(
        select(BUltraResult).where(BUltraResult.patient_id == patient_id)
    )
    gt = gt_result.scalar_one_or_none()

    asr_result = await db.execute(
        select(PatientAsrResult).where(PatientAsrResult.patient_id == patient_id)
    )
    asr_map = {a.id: a for a in asr_result.scalars().all()}

    tmpl_result = await db.execute(select(PromptTemplate))
    tmpl_map = {t.id: t for t in tmpl_result.scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "LLM历史"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "LLM结果ID", "检查记录ID", "病历号", "检查日期", "执行时间", "状态", "准确率",
        "ASR结果ID", "ASR模型名称", "ASR转写来源", "LLM模型名称",
        "提示词模板ID", "提示词模板名称", "提示词长度",
        "提示词内容", "ASR原始转写", "实际送入LLM的ASR文本",
        "LLM_右侧卵泡总数", "LLM_右侧卵泡明细", "LLM_左侧卵泡总数", "LLM_左侧卵泡明细",
        "LLM_内膜厚度", "LLM_内膜类型",
        "LLM_右卵巢长", "LLM_右卵巢宽", "LLM_左卵巢长", "LLM_左卵巢宽",
        "LLM_备注", "LLM_总结", "LLM_不确定内容", "LLM_原始返回",
        "真实_右侧卵泡总数", "真实_右侧卵泡明细", "真实_左侧卵泡总数", "真实_左侧卵泡明细",
        "真实_内膜厚度", "真实_内膜类型",
        "真实_右卵巢长", "真实_右卵巢宽", "真实_左卵巢长", "真实_左卵巢宽",
        "真实_备注",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_wrap
        cell.border = thin_border

    def follicles_to_str(follicles):
        if not follicles or not isinstance(follicles, list):
            return "-"
        return "; ".join(f"{f.get('size', '?')}x{f.get('count', '?')}" for f in follicles)

    # 真实 B 超
    gt_right_total = gt.right_follicle_total if gt else ""
    gt_left_total = gt.left_follicle_total if gt else ""
    gt_endo_thick = gt.endometrium_thickness if gt else ""
    gt_endo_type = gt.endometrium_type if gt else ""
    gt_r_ovary_l = gt.right_ovary_length if gt else ""
    gt_r_ovary_w = gt.right_ovary_width if gt else ""
    gt_l_ovary_l = gt.left_ovary_length if gt else ""
    gt_l_ovary_w = gt.left_ovary_width if gt else ""
    gt_remark = gt.remark if gt else ""

    for row_idx, (llm, _patient, _df) in enumerate(rows, 2):
        structured = normalize_structured_result(llm.structured_result or {})
        right_follicles = structured.get("right_follicles") or []
        left_follicles = structured.get("left_follicles") or []

        asr = asr_map.get(llm.asr_result_id) if llm.asr_result_id else None
        asr_model_name = asr.asr_model_name if asr else ""
        asr_transcript = asr.full_transcript if asr else ""

        tmpl_id = llm.prompt_template_id
        tmpl_name = llm.prompt_template_name or ""
        if not tmpl_name and tmpl_id and tmpl_id in tmpl_map:
            tmpl_name = tmpl_map[tmpl_id].name or ""

        row_data = [
            llm.id, patient_id, record_id, date_str,
            llm.created_at.strftime("%Y-%m-%d %H:%M:%S") if llm.created_at else "",
            llm.status, llm.accuracy,
            llm.asr_result_id, asr_model_name, "original", llm.llm_model_name,
            tmpl_id or "", tmpl_name or "未记录模板名称",
            len(llm.prompt_content) if llm.prompt_content else 0,
            llm.prompt_content or "", asr_transcript, asr_transcript,
            structured.get("right_follicle_total", ""), follicles_to_str(right_follicles),
            structured.get("left_follicle_total", ""), follicles_to_str(left_follicles),
            structured.get("endometrium_thickness", ""), structured.get("endometrium_type", ""),
            structured.get("right_ovary_length", ""), structured.get("right_ovary_width", ""),
            structured.get("left_ovary_length", ""), structured.get("left_ovary_width", ""),
            structured.get("remark", ""), llm.summary_text or "",
            structured.get("uncertain_text", ""), llm.raw_output or "",
            gt_right_total, follicles_to_str(gt.right_follicles if gt else []),
            gt_left_total, follicles_to_str(gt.left_follicles if gt else []),
            gt_endo_thick, gt_endo_type,
            gt_r_ovary_l, gt_r_ovary_w, gt_l_ovary_l, gt_l_ovary_w, gt_remark,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = [
        10, 12, 12, 10, 18, 8, 8,
        10, 15, 12, 15, 12, 15, 10, 50, 50, 50,
        12, 25, 12, 25, 10, 10, 10, 10, 10, 10,
        30, 40, 30, 50,
        12, 25, 12, 25, 10, 10, 10, 10, 10, 10, 30,
    ]
    for idx, w in enumerate(col_widths, 1):
        col_letter = chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # ASCII-only filename to avoid latin-1 encoding error
    safe_record = record_id.encode('ascii', 'ignore').decode()
    filename = f"LLM_history_{safe_record}_{date_str}.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
