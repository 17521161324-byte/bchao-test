"""
结果管理路由 - xlsx 上传与解析、真实 B 超结果读写
"""
import os
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.database import get_db
from app.models import DateFolder, PatientRecord, BUltraResult
from app.schemas import BUltraResultOut
from app.services.parser import parse_follicle_string, normalize_follicles
from app.config import settings

router = APIRouter()


def _follicle_total(follicles: list[dict]) -> int:
    return sum(int(f.get("count") or 0) for f in follicles)


def _normalize_b_ultra_obj(obj: BUltraResult | None) -> BUltraResult | None:
    """兼容历史数据：接口返回前确保卵泡字段是结构化数组。"""
    if not obj:
        return obj
    obj.right_follicles = parse_follicle_string(obj.right_follicles)
    obj.left_follicles = parse_follicle_string(obj.left_follicles)
    obj.right_follicle_total = _follicle_total(obj.right_follicles)
    obj.left_follicle_total = _follicle_total(obj.left_follicles)
    return obj


@router.post("/upload")
async def upload_result_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """上传 B 超结果 xlsx 文件"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="只支持 xlsx/xls 文件")

    save_path = os.path.join(settings.UPLOAD_DIR, file.filename)
    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)

    try:
        parsed = await db.execute(select(DateFolder))
        date_folders = {d.date: d for d in parsed.scalars().all()}
        count = parse_xlsx_to_db(save_path, date_folders, db)
        await db.commit()
        return {"message": "上传成功", "imported": count, "filename": file.filename}
    except Exception as e:
        import logging
        logging.error(f"xlsx 解析失败: {e}")
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")


# ======= 历史兼容接口（按病历号）=======

@router.get("/{record_id}", response_model=BUltraResultOut | None)
async def get_result_by_record(record_id: str, db: AsyncSession = Depends(get_db)):
    """根据病历号获取结果（历史兼容）"""
    result = await db.execute(
        select(BUltraResult).where(BUltraResult.record_id == record_id)
    )
    return _normalize_b_ultra_obj(result.scalar_one_or_none())


@router.put("/{result_id}")
async def update_result(result_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """更新 B 超结果（历史兼容，按 result_id）"""
    result = await db.execute(
        select(BUltraResult).where(BUltraResult.id == result_id)
    )
    obj = result.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="结果不存在")

    if "right_follicles" in data:
        obj.right_follicles = parse_follicle_string(data["right_follicles"])
        obj.right_follicle_total = _follicle_total(obj.right_follicles)
    if "left_follicles" in data:
        obj.left_follicles = parse_follicle_string(data["left_follicles"])
        obj.left_follicle_total = _follicle_total(obj.left_follicles)

    field_map = {
        "endometrium_thickness": "endometrium_thickness",
        "endometrium_type": "endometrium_type",
        "right_ovary_length": "right_ovary_length",
        "right_ovary_width": "right_ovary_width",
        "left_ovary_length": "left_ovary_length",
        "left_ovary_width": "left_ovary_width",
        "remark": "remark",
    }
    for key, field in field_map.items():
        if key in data:
            setattr(obj, field, data[key])

    await db.commit()
    await db.refresh(obj)
    return _normalize_b_ultra_obj(obj)


# ======= 按检查记录 ID（推荐接口）=======

@router.get("/exam/{exam_record_id}/b-ultra")
async def get_b_ultra_by_exam(exam_record_id: int, db: AsyncSession = Depends(get_db)):
    """获取指定检查记录的真实 B 超结果"""
    patient = await db.get(PatientRecord, exam_record_id)
    if not patient:
        raise HTTPException(status_code=404, detail="检查记录不存在")

    result = await db.execute(
        select(BUltraResult).where(BUltraResult.patient_id == exam_record_id)
    )
    return _normalize_b_ultra_obj(result.scalar_one_or_none())


@router.put("/exam/{exam_record_id}/b-ultra")
async def upsert_b_ultra_by_exam(exam_record_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """新建或更新指定检查记录的真实 B 超结果"""
    patient = await db.get(PatientRecord, exam_record_id)
    if not patient:
        raise HTTPException(status_code=404, detail="检查记录不存在")

    result = await db.execute(
        select(BUltraResult).where(BUltraResult.patient_id == exam_record_id)
    )
    obj = result.scalar_one_or_none()

    right_source = data["right_follicles"] if "right_follicles" in data else (obj.right_follicles if obj else [])
    left_source = data["left_follicles"] if "left_follicles" in data else (obj.left_follicles if obj else [])
    right_follicles = normalize_follicles(parse_follicle_string(right_source))
    left_follicles = normalize_follicles(parse_follicle_string(left_source))

    r_total = data.get("right_follicle_total")
    l_total = data.get("left_follicle_total")
    if right_follicles:
        r_total = _follicle_total(right_follicles)
    elif r_total is None:
        r_total = 0
    if left_follicles:
        l_total = _follicle_total(left_follicles)
    elif l_total is None:
        l_total = 0

    field_map = {
        "endometrium_thickness": "endometrium_thickness",
        "endometrium_type": "endometrium_type",
        "right_ovary_length": "right_ovary_length",
        "right_ovary_width": "right_ovary_width",
        "left_ovary_length": "left_ovary_length",
        "left_ovary_width": "left_ovary_width",
        "remark": "remark",
    }

    update_data = {
        "right_follicles": right_follicles,
        "left_follicles": left_follicles,
        "right_follicle_total": r_total,
        "left_follicle_total": l_total,
    }
    for key, field in field_map.items():
        if key in data:
            update_data[field] = data[key]

    if obj:
        for k, v in update_data.items():
            setattr(obj, k, v)
    else:
        obj = BUltraResult(
            patient_id=exam_record_id,
            record_id=patient.record_id,
            date=patient.date_folder.date if patient.date_folder else "",
            source_file="manual",
            **update_data,
        )
        db.add(obj)

    await db.commit()
    await db.refresh(obj)
    return _normalize_b_ultra_obj(obj)


async def parse_xlsx_to_db(filepath: str, date_folders: dict, db: AsyncSession) -> int:
    """解析 xlsx 文件并写入数据库"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    if ws.max_row < 2:
        return 0

    # 获取表头
    headers = [cell.value for cell in ws[1]]
    header_map = {str(h).strip(): i for i, h in enumerate(headers) if h}

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record_id = str(row[0]).strip()

        # 查找对应的 patient
        date_str = infer_date_from_filename(os.path.basename(filepath))
        date_folder = date_folders.get(date_str)

        if not date_folder:
            # 尝试创建日期文件夹
            date_folder = DateFolder(date=date_str, path="")
            db.add(date_folder)
            await db.flush()
            date_folders[date_str] = date_folder

        # 查找 patient
        from sqlalchemy import select as sa_select
        result = await db.execute(
            sa_select(PatientRecord).where(
                PatientRecord.record_id == record_id,
                PatientRecord.date_folder_id == date_folder.id,
            )
        )
        patient = result.scalar_one_or_none()
        if not patient:
            patient = PatientRecord(
                record_id=record_id,
                date_folder_id=date_folder.id,
                timestamp_folder="",
            )
            db.add(patient)
            await db.flush()

        # 解析各字段
        def get_val(field_names, default=None):
            for name in field_names:
                if name in header_map:
                    idx = header_map[name]
                    if idx < len(row):
                        return row[idx]
            return default

        right_follicle_str = get_val(["右侧卵泡数", "右卵泡数", "右卵泡"], "")
        left_follicle_str = get_val(["左侧卵泡数", "左卵泡数", "左卵泡"], "")

        right_follicles = parse_follicle_string(str(right_follicle_str))
        left_follicles = parse_follicle_string(str(left_follicle_str))

        result_obj = BUltraResult(
            patient_id=patient.id,
            record_id=record_id,
            date=date_str,
            source_file=os.path.basename(filepath),
            right_follicles=right_follicles,
            left_follicles=left_follicles,
            right_follicle_total=sum(f["count"] for f in right_follicles),
            left_follicle_total=sum(f["count"] for f in left_follicles),
            endometrium_thickness=safe_float(get_val(["内膜厚度"])),
            endometrium_type=safe_str(get_val(["内膜类型"])),
            right_ovary_length=safe_float(get_val(["右卵巢长", "右卵巢长度"])),
            right_ovary_width=safe_float(get_val(["右卵巢宽", "右卵巢宽度"])),
            left_ovary_length=safe_float(get_val(["左卵巢长", "左卵巢长度"])),
            left_ovary_width=safe_float(get_val(["左卵巢宽", "左卵巢宽度"])),
            remark=safe_str(get_val(["备注", "remark"], "")),
        )

        # 检查是否已存在
        existing = await db.execute(
            sa_select(BUltraResult).where(BUltraResult.patient_id == patient.id)
        )
        old = existing.scalar_one_or_none()
        if old:
            # 更新
            for k, v in result_obj.__dict__.items():
                if not k.startswith("_"):
                    setattr(old, k, v)
        else:
            db.add(result_obj)

        count += 1

    return count


def infer_date_from_filename(filename: str) -> str:
    """从文件名推断日期（如 20260623.xlsx → 20260623）"""
    import re
    m = re.search(r"(\d{8})", filename)
    return m.group(1) if m else "unknown"


def safe_float(val):
    if val is None or str(val).strip().upper() == "NULL":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val):
    if val is None or str(val).strip().upper() == "NULL":
        return None
    return str(val).strip()
