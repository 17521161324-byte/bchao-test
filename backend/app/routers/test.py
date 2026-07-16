"""
测试执行路由（含 SSE 流式进度）
"""
import asyncio
import json
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from loguru import logger

from app.config import resolve_hotwords
from app.database import get_db
from app.models import PatientRecord, AudioSeg, ModelConfig, TestRun, BUltraResult
from app.schemas import TestStartRequest, TestResultOut, EvaluationUpdate
from app.services.asr import create_asr
from app.services.asr_input import build_asr_audio_inputs
from app.services.test_executor import TestExecutor
from app.services.parser import evaluate_result

router = APIRouter()


@router.get("/history", response_model=list[TestResultOut])
async def list_test_history(
    record_id: str | None = None,
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
):
    """获取测试历史列表"""
    query = select(TestRun).order_by(TestRun.created_at.desc()).offset(skip).limit(limit)
    if record_id:
        query = query.where(TestRun.record_id == record_id)
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/asr")
async def run_asr_only(
    record_id: str = Query(...),
    asr_model_id: int = Query(...),
    hotwords: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    """单独运行 ASR（同步返回，不经过 SSE）"""
    result = await db.execute(
        select(PatientRecord)
        .options(selectinload(PatientRecord.segs))
        .where(PatientRecord.record_id == record_id)
        .order_by(PatientRecord.id.desc())
    )
    patients = result.scalars().all()
    if not patients:
        raise HTTPException(status_code=404, detail=f"病历号 {record_id} 不存在")
    patient = next((p for p in patients if p.segs), patients[0])

    if not patient.segs:
        raise HTTPException(status_code=400, detail="该病历无录音文件")

    asr_result = await db.execute(
        select(ModelConfig).where(ModelConfig.id == asr_model_id)
    )
    asr_model = asr_result.scalar_one_or_none()
    if not asr_model:
        raise HTTPException(status_code=404, detail="ASR 模型不存在")

    raw_segs = [
        {"seg_index": s.seg_index, "file_path": s.file_path, "duration": s.duration}
        for s in sorted(patient.segs, key=lambda x: x.seg_index)
    ]

    asr = create_asr(asr_model.provider, **{
        "endpoint": asr_model.endpoint,
        "api_key": asr_model.api_key,
        "api_secret": asr_model.api_secret,
        "secret_key": asr_model.secret_key,
        "model_name": asr_model.model_name,
        "params": asr_model.params or {},
    })

    # 按优先级解析热词: 接口 > 模型配置 > 默认
    parsed_hotwords = hotwords.split(",") if hotwords else None
    resolved_hotwords = resolve_hotwords(parsed_hotwords, asr_model.params)

    actual_segs = build_asr_audio_inputs(raw_segs, asr_model.params or {})
    asr_results = []
    for seg in actual_segs:
        text = await asr.transcribe(seg["file_path"], hotwords=resolved_hotwords)
        asr_results.append({
            "seg_index": seg["seg_index"],
            "text": text,
            "duration": seg["duration"],
            "input_mode": seg.get("input_mode", "segments"),
            "source_seg_count": seg.get("source_seg_count"),
        })

    return {
        "model_name": asr_model.name,
        "model_id": asr_model.id,
        "record_id": record_id,
        "segments": asr_results,
        "full_transcript": "\n".join(r["text"] for r in asr_results),
    }


@router.get("/asr/stream")
async def run_asr_stream(
    record_id: str = Query(...),
    asr_model_id: int = Query(...),
    hotwords: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    """单独运行 ASR (SSE 流式返回, 每完成一个 seg 即时推送)

    事件类型:
    - progress: 开始处理一个 seg, data = {"stage": "progress", "seg_index": int, "total": int}
    - segment: 一个 seg 转写完成, data = {"stage": "segment", "seg_index": int, "text": str, "duration": float}
    - complete: 全部完成, data = {"stage": "complete", "segments": [...], "full_transcript": str}
    - error: 出错, data = {"stage": "error", "message": str}
    """
    result = await db.execute(
        select(PatientRecord)
        .options(selectinload(PatientRecord.segs))
        .where(PatientRecord.record_id == record_id)
        .order_by(PatientRecord.id.desc())
    )
    patients = result.scalars().all()
    if not patients:
        raise HTTPException(status_code=404, detail=f"病历号 {record_id} 不存在")
    patient = next((p for p in patients if p.segs), patients[0])

    if not patient.segs:
        raise HTTPException(status_code=400, detail="该病历无录音文件")

    asr_result = await db.execute(
        select(ModelConfig).where(ModelConfig.id == asr_model_id)
    )
    asr_model = asr_result.scalar_one_or_none()
    if not asr_model:
        raise HTTPException(status_code=404, detail="ASR 模型不存在")

    raw_segs = [
        {"seg_index": s.seg_index, "file_path": s.file_path, "duration": s.duration}
        for s in sorted(patient.segs, key=lambda x: x.seg_index)
    ]

    asr = create_asr(asr_model.provider, **{
        "endpoint": asr_model.endpoint,
        "api_key": asr_model.api_key,
        "api_secret": asr_model.api_secret,
        "secret_key": asr_model.secret_key,
        "model_name": asr_model.model_name,
        "params": asr_model.params or {},
    })

    # 解析热词优先级: 接口 > 模型配置 > 默认
    parsed_hotwords = hotwords.split(",") if hotwords else None
    resolved_hotwords = resolve_hotwords(parsed_hotwords, asr_model.params)

    async def event_generator():
        asr_results = []
        actual_segs = build_asr_audio_inputs(raw_segs, asr_model.params or {})
        total = len(actual_segs)
        try:
            for seg in actual_segs:
                yield f"event: progress\ndata: {json.dumps({'stage': 'progress', 'seg_index': seg['seg_index'], 'total': total}, ensure_ascii=False)}\n\n"

                text = await asr.transcribe(seg["file_path"], hotwords=resolved_hotwords)
                asr_results.append({
                    "seg_index": seg["seg_index"],
                    "text": text,
                    "duration": seg["duration"],
                    "input_mode": seg.get("input_mode", "segments"),
                    "source_seg_count": seg.get("source_seg_count"),
                })

                yield f"event: segment\ndata: {json.dumps({'stage': 'segment', 'seg_index': seg['seg_index'], 'text': text, 'duration': seg['duration']}, ensure_ascii=False)}\n\n"

            full_transcript = "\n".join(r["text"] for r in asr_results)
            yield f"event: complete\ndata: {json.dumps({'stage': 'complete', 'segments': asr_results, 'full_transcript': full_transcript}, ensure_ascii=False)}\n\n"
        except Exception as e:
            logger.error(f"ASR 流式转写失败: {e}")
            yield f"event: error\ndata: {json.dumps({'stage': 'error', 'message': str(e)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/start")
async def start_test(
    record_id: str = Query(...),
    asr_model_id: int = Query(...),
    llm_model_id: int | None = Query(None),
    prompt_version: str = Query("v1.0"),
    prompt_template: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    """
    启动单条测试（返回 SSE 流实时推送进度）
    使用 GET 查询参数，兼容浏览器 EventSource
    """
    # 查找病历
    result = await db.execute(
        select(PatientRecord)
        .options(selectinload(PatientRecord.segs), selectinload(PatientRecord.date_folder))
        .where(PatientRecord.record_id == record_id)
        .order_by(PatientRecord.id)
    )
    patients = result.scalars().all()
    if not patients:
        raise HTTPException(status_code=404, detail=f"病历号 {record_id} 不存在")
    patient = patients[0]

    if not patient.segs:
        raise HTTPException(status_code=400, detail="该病历无录音文件")

    # 查找 ASR 模型
    asr_result = await db.execute(
        select(ModelConfig).where(ModelConfig.id == asr_model_id)
    )
    asr_model = asr_result.scalar_one_or_none()
    if not asr_model:
        raise HTTPException(status_code=404, detail="ASR 模型不存在")

    # 查找 LLM 模型（可选）
    llm_model = None
    if llm_model_id:
        llm_result = await db.execute(
            select(ModelConfig).where(ModelConfig.id == llm_model_id)
        )
        llm_model = llm_result.scalar_one_or_none()

    # 准备 seg 数据
    segs = [
        {
            "seg_index": s.seg_index,
            "file_path": s.file_path,
            "duration": s.duration,
        }
        for s in sorted(patient.segs, key=lambda x: x.seg_index)
    ]

    # 创建测试记录
    from app.models import DateFolder
    date_str = ""
    if patient.date_folder:
        date_str = patient.date_folder.date

    test_run = TestRun(
        record_id=record_id,
        date=date_str,
        asr_model_id=asr_model.id,
        llm_model_id=llm_model.id if llm_model else None,
        prompt_version=prompt_version,
        asr_results=[],
        full_transcript="",
    )
    db.add(test_run)
    await db.commit()
    await db.refresh(test_run)

    async def event_generator():
        executor = TestExecutor()
        queue = asyncio.Queue()

        async def on_progress(event):
            await queue.put(event)

        # 构造配置
        asr_config = {
            "endpoint": asr_model.endpoint,
            "api_key": asr_model.api_key,
            "api_secret": asr_model.api_secret,
            "secret_key": asr_model.secret_key,
            "model_name": asr_model.model_name,
            "params": asr_model.params or {},
        }
        llm_config = None
        if llm_model:
            llm_config = {
                "endpoint": llm_model.endpoint,
                "api_key": llm_model.api_key,
                "api_secret": llm_model.api_secret,
                "model_name": llm_model.model_name,
            }

        async def run_test():
            try:
                result_data = await executor.execute(
                    segs=segs,
                    asr_provider=asr_model.provider,
                    asr_config=asr_config,
                    llm_provider=llm_model.provider if llm_model else None,
                    llm_config=llm_config,
                    prompt_template=prompt_template,
                    progress_callback=on_progress,
                )

                # 更新测试记录
                test_run.asr_results = result_data["asr_results"]
                test_run.full_transcript = result_data["full_transcript"]
                test_run.llm_raw_output = result_data["llm_raw_output"]
                test_run.structured_result = result_data["structured_result"]
                test_run.summary_text = result_data["summary_text"]
                test_run.duration_seconds = result_data["duration_seconds"]

                # 自动评估（如果有 ground truth）
                gt_result = await db.execute(
                    select(BUltraResult).where(BUltraResult.patient_id == patient.id)
                )
                gt = gt_result.scalar_one_or_none()
                if gt and test_run.structured_result:
                    evaluation = evaluate_result(
                        identified=test_run.structured_result,
                        ground_truth={
                            "right_follicle_total": gt.right_follicle_total,
                            "left_follicle_total": gt.left_follicle_total,
                            "endometrium_thickness": gt.endometrium_thickness,
                            "endometrium_type": gt.endometrium_type,
                            "right_ovary_length": gt.right_ovary_length,
                            "right_ovary_width": gt.right_ovary_width,
                            "left_ovary_length": gt.left_ovary_length,
                            "left_ovary_width": gt.left_ovary_width,
                        }
                    )
                    test_run.evaluation = evaluation
                    test_run.accuracy = evaluation.get("accuracy")

                await db.commit()
                await queue.put({"stage": "complete", "test_id": test_run.id, "status": "ok"})
            except Exception as e:
                logger.error(f"测试执行失败: {e}")
                await queue.put({"stage": "error", "message": str(e)})
            finally:
                await queue.put(None)  # Sentinel to stop generator

        # Start test execution in background
        task = asyncio.create_task(run_test())

        try:
            while True:
                event = await asyncio.wait_for(queue.get(), timeout=300)
                if event is None:
                    break
                if event["stage"] == "complete":
                    yield f"event: complete\ndata: {json.dumps(event, ensure_ascii=False)}\n\n"
                    break
                elif event["stage"] == "error":
                    yield f"event: error\ndata: {json.dumps(event, ensure_ascii=False)}\n\n"
                    break
                else:
                    yield f"event: progress\ndata: {json.dumps(event, ensure_ascii=False)}\n\n"
        finally:
            if not task.done():
                task.cancel()

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/{test_id}", response_model=TestResultOut)
async def get_test_result(test_id: int, db: AsyncSession = Depends(get_db)):
    """获取测试结果"""
    result = await db.execute(select(TestRun).where(TestRun.id == test_id))
    test = result.scalar_one_or_none()
    if not test:
        raise HTTPException(status_code=404, detail="测试记录不存在")
    return test


@router.post("/llm")
async def run_llm_extraction(data: dict, db: AsyncSession = Depends(get_db)):
    """LLM 结构化提取（给定转写文本）"""
    from app.services.llm import create_llm
    from app.services.parser import DEFAULT_PROMPT_TEMPLATE

    transcript = data.get("transcript", "")
    if not transcript:
        raise HTTPException(status_code=400, detail="转写文本为空")

    llm_model_id = data.get("llm_model_id")
    prompt_template = data.get("prompt_template", DEFAULT_PROMPT_TEMPLATE)

    llm_model = None
    if llm_model_id:
        result = await db.execute(select(ModelConfig).where(ModelConfig.id == llm_model_id))
        llm_model = result.scalar_one_or_none()

    if llm_model:
        llm = create_llm(llm_model.provider, **{
            "endpoint": llm_model.endpoint,
            "api_key": llm_model.api_key,
            "api_secret": llm_model.api_secret,
            "model_name": llm_model.model_name,
        })
    else:
        raise HTTPException(status_code=400, detail="请提供有效的 LLM 模型")

    try:
        response = await llm.extract(transcript, prompt_template)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LLM 调用失败: {str(e)}")

    return {
        "model_name": llm_model.name if llm_model else "unknown",
        "model_id": llm_model.id if llm_model else None,
        "structured": response.structured,
        "summary": response.summary,
    }


@router.put("/{test_id}/evaluate", response_model=TestResultOut)
async def update_evaluation(
    test_id: int,
    data: EvaluationUpdate,
    db: AsyncSession = Depends(get_db),
):
    """提交人工修正评估"""
    result = await db.execute(select(TestRun).where(TestRun.id == test_id))
    test = result.scalar_one_or_none()
    if not test:
        raise HTTPException(status_code=404, detail="测试记录不存在")

    test.structured_result = data.structured_result
    test.human_corrected = data.human_corrected

    # 重新评估
    patient = await db.execute(
        select(PatientRecord).where(PatientRecord.record_id == test.record_id)
    )
    patient_obj = patient.scalar_one_or_none()
    if patient_obj:
        gt_result = await db.execute(
            select(BUltraResult).where(BUltraResult.patient_id == patient_obj.id)
        )
        gt = gt_result.scalar_one_or_none()
        if gt:
            evaluation = evaluate_result(
                identified=data.structured_result,
                ground_truth={
                    "right_follicle_total": gt.right_follicle_total,
                    "left_follicle_total": gt.left_follicle_total,
                    "endometrium_thickness": gt.endometrium_thickness,
                    "endometrium_type": gt.endometrium_type,
                    "right_ovary_length": gt.right_ovary_length,
                    "right_ovary_width": gt.right_ovary_width,
                    "left_ovary_length": gt.left_ovary_length,
                    "left_ovary_width": gt.left_ovary_width,
                }
            )
            test.evaluation = evaluation
            test.accuracy = evaluation.get("accuracy")

    await db.commit()
    await db.refresh(test)
    return test


# ------------------------------------------------------------------
# LLM 历史记录 (用于批量查看和导出)
# ------------------------------------------------------------------

from fastapi.responses import StreamingResponse as SSEStreamingResponse
import io
from datetime import datetime
from app.models import PatientLlmResult, PatientAsrResult


@router.get("/llm-history")
async def list_llm_history(
    record_id: Optional[str] = None,
    llm_model_name: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
):
    """查询 LLM 历史记录 (跨患者)"""
    query = (
        select(PatientLlmResult, PatientRecord)
        .join(PatientRecord, PatientLlmResult.patient_id == PatientRecord.id)
        .order_by(PatientLlmResult.created_at.desc())
        .offset(skip)
        .limit(limit)
    )

    if record_id:
        query = query.where(PatientRecord.record_id == record_id)
    if llm_model_name:
        query = query.where(PatientLlmResult.llm_model_name == llm_model_name)
    if status:
        query = query.where(PatientLlmResult.status == status)
    if date_from:
        query = query.where(PatientLlmResult.created_at >= date_from)
    if date_to:
        query = query.where(PatientLlmResult.created_at <= date_to)

    result = await db.execute(query)
    rows = result.all()

    output = []
    for llm, patient in rows:
        output.append({
            "id": llm.id,
            "exam_record_id": patient.id,
            "record_id": patient.record_id,
            "date": patient.date_folder.date if patient.date_folder else None,
            "asr_model_name": None,
            "llm_model_name": llm.llm_model_name,
            "prompt_version": llm.prompt_version,
            "prompt_len": len(llm.prompt_content) if llm.prompt_content else 0,
            "summary_text": llm.summary_text,
            "structured_result": llm.structured_result,
            "accuracy": llm.accuracy,
            "status": llm.status,
            "error_message": llm.error_message,
            "created_at": llm.created_at.isoformat() if llm.created_at else None,
        })
    return output


@router.get("/llm-history/export")
async def export_llm_history(
    record_id: Optional[str] = None,
    llm_model_name: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """导出 LLM 历史为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = (
        select(PatientLlmResult, PatientRecord)
        .join(PatientRecord, PatientLlmResult.patient_id == PatientRecord.id)
        .order_by(PatientLlmResult.created_at.desc())
        .limit(5000)
    )
    if record_id:
        query = query.where(PatientRecord.record_id == record_id)
    if llm_model_name:
        query = query.where(PatientLlmResult.llm_model_name == llm_model_name)
    if status:
        query = query.where(PatientLlmResult.status == status)
    if date_from:
        query = query.where(PatientLlmResult.created_at >= date_from)
    if date_to:
        query = query.where(PatientLlmResult.created_at <= date_to)

    result = await db.execute(query)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "LLM历史"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "ID", "病历号", "日期", "检查记录ID",
        "ASR模型", "LLM模型", "提示词版本", "提示词长度",
        "状态", "准确率", "内膜厚度", "内膜类型",
        "右卵巢长", "右卵巢宽", "左卵巢长", "左卵巢宽",
        "右卵泡总数", "左卵泡总数", "备注", "总结",
        "创建时间",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (llm, patient) in enumerate(rows, 2):
        structured = llm.structured_result or {}
        row_data = [
            llm.id,
            patient.record_id,
            patient.date_folder.date if patient.date_folder else "",
            patient.id,
            llm.asr_result.asr_model_name if llm.asr_result else "",
            llm.llm_model_name,
            llm.prompt_version,
            len(llm.prompt_content) if llm.prompt_content else 0,
            llm.status,
            llm.accuracy,
            structured.get("endometrium_thickness", ""),
            structured.get("endometrium_type", ""),
            structured.get("right_ovary_length", ""),
            structured.get("right_ovary_width", ""),
            structured.get("left_ovary_length", ""),
            structured.get("left_ovary_width", ""),
            structured.get("right_follicle_total", ""),
            structured.get("left_follicle_total", ""),
            structured.get("remark", ""),
            llm.summary_text or "",
            llm.created_at.strftime("%Y-%m-%d %H:%M:%S") if llm.created_at else "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = [6, 12, 10, 12, 15, 15, 10, 10, 8, 8, 10, 10, 10, 10, 10, 10, 10, 10, 30, 50, 20]
    for idx, w in enumerate(col_widths, 1):
        col_letter = chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)
        ws.column_dimensions[col_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    return SSEStreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=llm_history_{timestamp}.xlsx"},
    )
