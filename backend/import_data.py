"""
数据导入脚本：扫描录音目录 + 解析 xlsx 结果文件 → 写入数据库
"""
import os
import re
import glob
import asyncio
import openpyxl
from sqlalchemy import select
from app.database import AsyncSessionLocal, init_db
from app.models import DateFolder, PatientRecord, AudioSeg, BUltraResult


RECORDINGS_DIR = r"E:\bchao-test\backend\data\recordings"
RESULTS_DIR = r"E:\测试平台\B超结果文件"

# 中文键名 → 英文字段名（兼容不同表格的表头名称）
HEADER_MAP = {
    "病历号": "record_id",
    "右侧卵泡数": "right_follicle_str",
    "右卵泡数": "right_follicle_str",
    "右卵泡": "right_follicle_str",
    "左侧卵泡数": "left_follicle_str",
    "左卵泡数": "left_follicle_str",
    "左卵泡": "left_follicle_str",
    "备注": "remark",
    "remark": "remark",
    "内膜厚度": "endometrium_thickness",
    "内膜类型": "endometrium_type",
    "右卵巢长": "right_ovary_length",
    "右卵巢长度": "right_ovary_length",
    "右卵巢宽": "right_ovary_width",
    "右卵巢宽度": "right_ovary_width",
    "左卵巢长": "left_ovary_length",
    "左卵巢长度": "left_ovary_length",
    "左卵巢宽": "left_ovary_width",
    "左卵巢宽度": "left_ovary_width",
}


def parse_follicle_string(s: str) -> list:
    if not s or str(s).strip().upper() == "NULL":
        return []
    result = []
    pattern = r"(\d+\.?\d*)\s*[×﹡*]\s*(\d+)"
    for size_str, count_str in re.findall(pattern, str(s)):
        try:
            result.append({"size": float(size_str), "count": int(count_str)})
        except ValueError:
            continue
    return result


def safe_float(val):
    if val is None or str(val).strip().upper() == "NULL":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val):
    if val is None or str(val).strip().upper() == "NULL":
        return None
    return str(val).strip()


def infer_date_from_filename(filename: str) -> str:
    m = re.search(r"(\d{8})", filename)
    return m.group(1) if m else "unknown"


async def import_recordings():
    """扫描录音目录，写入日期→病历号→分段"""
    async with AsyncSessionLocal() as db:
        date_folders = {}
        total_segs = 0

        for date_dir in sorted(os.listdir(RECORDINGS_DIR)):
            date_path = os.path.join(RECORDINGS_DIR, date_dir)
            if not os.path.isdir(date_path):
                continue
            if date_dir.startswith("."):
                continue

            # 查找或创建 DateFolder
            result = await db.execute(select(DateFolder).where(DateFolder.date == date_dir))
            df = result.scalar_one_or_none()
            if not df:
                df = DateFolder(date=date_dir, path=date_path)
                db.add(df)
                await db.flush()
            date_folders[date_dir] = df

            for record_id in sorted(os.listdir(date_path)):
                record_path = os.path.join(date_path, record_id)
                if not os.path.isdir(record_path):
                    continue
                if record_id.startswith("."):
                    continue

                # 查找或创建 PatientRecord
                result = await db.execute(
                    select(PatientRecord).where(
                        PatientRecord.record_id == record_id,
                        PatientRecord.date_folder_id == df.id,
                    )
                )
                patient = result.scalar_one_or_none()
                if not patient:
                    patient = PatientRecord(
                        record_id=record_id,
                        date_folder_id=df.id,
                        timestamp_folder="",
                    )
                    db.add(patient)
                    await db.flush()

                # 遍历时间戳文件夹
                for ts_folder in os.listdir(record_path):
                    ts_path = os.path.join(record_path, ts_folder)
                    if not os.path.isdir(ts_path):
                        continue
                    audio_path = os.path.join(ts_path, "audio")
                    if not os.path.isdir(audio_path):
                        continue

                    # 如果 timestamp_folder 为空，更新它
                    if not patient.timestamp_folder:
                        patient.timestamp_folder = ts_folder

                    # 扫描 wav 文件
                    for fname in sorted(os.listdir(audio_path)):
                        if not fname.endswith(".wav"):
                            continue
                        fpath = os.path.join(audio_path, fname)
                        file_size = os.path.getsize(fpath)

                        # 检查是否已存在
                        seg_idx = int(re.search(r"seg-(\d+)", fname).group(1))
                        existing = await db.execute(
                            select(AudioSeg).where(
                                AudioSeg.patient_id == patient.id,
                                AudioSeg.seg_index == seg_idx,
                            )
                        )
                        if existing.scalar_one_or_none():
                            continue

                        seg = AudioSeg(
                            patient_id=patient.id,
                            seg_index=seg_idx,
                            filename=fname,
                            file_path=fpath,
                            file_size=file_size,
                        )
                        db.add(seg)
                        total_segs += 1

        await db.commit()
        print(f"  [OK] recordings: {len(date_folders)} dates, {total_segs} segs")


async def import_results():
    """解析 xlsx 结果文件"""
    async with AsyncSessionLocal() as db:
        xlsx_files = glob.glob(os.path.join(RESULTS_DIR, "*.xlsx"))
        total = 0

        for filepath in sorted(xlsx_files):
            fname = os.path.basename(filepath)
            date_str = infer_date_from_filename(fname)
            print(f"  Parsing {fname} -> date {date_str}")

            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            if ws.max_row < 2:
                print(f"    ! empty, skip")
                continue

            headers = [cell.value for cell in ws[1]]
            header_map = {str(h).strip(): i for i, h in enumerate(headers) if h}

            # 查找 DateFolder
            result = await db.execute(select(DateFolder).where(DateFolder.date == date_str))
            df = result.scalar_one_or_none()
            if not df:
                df = DateFolder(date=date_str, path="")
                db.add(df)
                await db.flush()

            count = 0
            async with db.begin_nested():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue

                    record_id = str(row[0]).strip()

                    # 查找 PatientRecord
                    result = await db.execute(
                        select(PatientRecord).where(
                            PatientRecord.record_id == record_id,
                            PatientRecord.date_folder_id == df.id,
                        )
                    )
                    patient = result.scalar_one_or_none()
                    if not patient:
                        patient = PatientRecord(
                            record_id=record_id,
                            date_folder_id=df.id,
                            timestamp_folder="",
                        )
                        db.add(patient)
                        await db.flush()

                    def get_val(field_names, default=None):
                        for name in field_names:
                            en = HEADER_MAP.get(name)
                            if en and name in header_map:
                                idx = header_map[name]
                                if idx < len(row):
                                    return row[idx]
                        return default

                    right_follicles = parse_follicle_string(str(get_val(["右侧卵泡数", "右卵泡数", "右卵泡"], "")))
                    left_follicles = parse_follicle_string(str(get_val(["左侧卵泡数", "左卵泡数", "左卵泡"], "")))

                    result_obj = BUltraResult(
                        patient_id=patient.id,
                        record_id=record_id,
                        date=date_str,
                        source_file=fname,
                        right_follicles=right_follicles,
                        left_follicles=left_follicles,
                        right_follicle_total=sum(f["count"] for f in right_follicles),
                        left_follicle_total=sum(f["count"] for f in left_follicles),
                        endometrium_thickness=safe_float(get_val(["内膜厚度"])),
                        endometrium_type=safe_str(get_val(["内膜类型"])),
                        right_ovary_length=safe_float(get_val(["右卵巢长", "右卵巢长度"])),
                        right_ovary_width=safe_float(get_val(["右卵巢宽", "右卵巢宽度"])),
                        left_ovary_length=safe_float(get_val(["左卵巢长", "左卵巢长度"])),
                        left_ovary_width=safe_float(get_val(["左卵巢宽", "左卵巢宽度"])),
                        remark=safe_str(get_val(["备注", "remark"])),
                    )

                    # upsert
                    existing = await db.execute(
                        select(BUltraResult).where(BUltraResult.patient_id == patient.id)
                    )
                    old = existing.scalar_one_or_none()
                    if old:
                        for k, v in result_obj.__dict__.items():
                            if not k.startswith("_"):
                                setattr(old, k, v)
                    else:
                        db.add(result_obj)

                    count += 1
                    print(f"    -> {record_id}: R={result_obj.right_follicle_total} L={result_obj.left_follicle_total}")

                await db.commit()
                total += count
                print(f"    [OK] imported {count} rows")

        print(f"  [OK] results: total {total} rows")


async def main():
    print("=" * 50)
    print("  Data Import Script")
    print("=" * 50)

    print("\n[1/3] Init db...")
    await init_db()

    print("\n[2/3] Skipping recordings (already imported)")

    print("\n[3/3] Import results...")
    await import_results()

    print("\n" + "=" * 50)
    print("  Import done!")
    print("=" * 50)


if __name__ == "__main__":
    asyncio.run(main())
